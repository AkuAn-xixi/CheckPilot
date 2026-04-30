import openpyxl
import re
import json
from pathlib import Path
from collections import Counter

# 默认合法按键名称（当配置文件不存在时使用）
_DEFAULT_VALID_KEYS = {
    'OK', 'RIGHT', 'UP', 'LEFT', 'DOWN', 'SETTING', 'HOME', 'POWER', 'BACK',
    'SOURCE', 'MENU', 'CHUP', 'CHDOWN', 'DIGITAL', 'EXITMENU', 'DIGITAL1',
    'DIGITAL2', 'DIGITAL3', 'DIGITAL4', 'DIGITAL5', 'DIGITAL6', 'DIGITAL7',
    'DIGITAL8', 'DIGITAL9', 'DIGITAL0', 'LIBRARY', 'TV_AV', 'VOLUMEUP',
    'VOLUMEDOWN', 'NETFLIX', 'YOUTUBE', 'PRIME_VIDEO', 'ACTION3', 'APPS', 'FILES', 'MUTE'
}

_CUSTOMIZATION_FILE = Path(__file__).resolve().parents[1] / "customization.json"


def get_valid_keys() -> set:
    """从配置文件加载合法按键集合，读取当前激活方案，不存在时返回默认值。"""
    if _CUSTOMIZATION_FILE.exists():
        try:
            with open(_CUSTOMIZATION_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            # 同时兼容旧扁平格式和新方案格式
            if "schemes" in data:
                active = data.get("active_scheme", "默认")
                scheme = data.get("schemes", {}).get(active, {})
                keys = scheme.get("valid_keys")
            else:
                keys = data.get("valid_keys")
            if isinstance(keys, list) and keys:
                return set(keys)
        except Exception:
            pass
    return set(_DEFAULT_VALID_KEYS)


# 定义所有合法的按键名称
VALID_KEYS = get_valid_keys()


def validate_excel(file_path):
    valid_keys = get_valid_keys()
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
    except Exception as e:
        print(f"无法打开Excel文件: {e}")
        return

    # 验证C列值是否唯一
    c_values = [str(sheet.cell(row=i, column=3).value) for i in range(2, sheet.max_row + 1)
                if sheet.cell(row=i, column=3).value is not None]

    if len(c_values) != len(set(c_values)):
        duplicates = [item for item, count in Counter(c_values).items() if count > 1]
        print(f"错误：C列值不唯一，重复的值有: {', '.join(duplicates)}")
    else:
        print("C列值验证通过：所有值都是唯一的")

    # 验证F和G列格式（按键名/次数/时间，多个用逗号分隔）
    for col in ['F', 'G']:
        col_idx = openpyxl.utils.column_index_from_string(col)
        for row in range(2, sheet.max_row + 1):
            cell_value = str(sheet.cell(row=row, column=col_idx).value)
            if cell_value == 'None' or not cell_value.strip():
                continue

            # 检查每个命令段
            commands = [cmd.strip() for cmd in cell_value.split(',') if cmd.strip()]
            for cmd in commands:
                parts = cmd.split('/')
                if len(parts) != 3:
                    print(f"错误：{col}{row} 命令'{cmd}'格式应为 KEY/COUNT/TIME")
                    continue

                key, count, time = parts
                if key not in valid_keys:
                    print(f"错误：{col}{row} 按键名称'{key}'无效")

                if not count.isdigit() or int(count) <= 0:
                    print(f"错误：{col}{row} 按键次数'{count}'必须为正整数")

                try:
                    float(time)
                except ValueError:
                    print(f"错误：{col}{row} 时间参数'{time}'必须为数字")

    # 验证I列格式（*.png）
    for row in range(2, sheet.max_row + 1):
        cell_value = str(sheet.cell(row=row, column=9).value)
        if cell_value == 'None' or not cell_value.strip():
            continue

        if not cell_value.lower().endswith('.png'):
            print(f"错误：I{row} 必须为PNG文件格式，当前值: {cell_value}")

    # 验证J列格式（(数字,数字)）使用英文括号和逗号
    for row in range(2, sheet.max_row + 1):
        cell_value = str(sheet.cell(row=row, column=10).value)
        if cell_value == 'None' or not cell_value.strip():
            continue

        # 检查英文括号和逗号
        if not (cell_value.startswith('(') and cell_value.endswith(')') and ',' in cell_value):
            print(f"错误：J{row} 必须使用英文括号和逗号，格式应为(数字,数字)，当前值: {cell_value}")
            continue

        # 提取数字部分
        try:
            num_part = cell_value[1:-1]  # 去掉括号
            num1, num2 = [num.strip() for num in num_part.split(',', 1)]
            float(num1)
            float(num2)
        except ValueError:
            print(f"错误：J{row} 包含非数字内容，当前值: {cell_value}")

    print("验证完成")


if __name__ == "__main__":
    file_path = input("请输入Excel文件路径: ")
    validate_excel(file_path)