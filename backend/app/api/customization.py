"""客制化配置 API — 支持多方案（scheme）管理"""
import copy
import io
import json
import re
from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel
from typing import Dict, List, Optional, Tuple

from ..config import settings
from ..utils.adb_controller import KEYCODE_MAP

router = APIRouter(prefix="/api/customization", tags=["customization"])

DEFAULT_SCHEME_NAME = "默认"

DEFAULT_VALID_KEYS: List[str] = sorted([
    'OK', 'RIGHT', 'UP', 'LEFT', 'DOWN', 'SETTING', 'HOME', 'POWER', 'BACK',
    'SOURCE', 'MENU', 'CHUP', 'CHDOWN', 'DIGITAL', 'EXITMENU', 'DIGITAL0',
    'DIGITAL1', 'DIGITAL2', 'DIGITAL3', 'DIGITAL4', 'DIGITAL5', 'DIGITAL6',
    'DIGITAL7', 'DIGITAL8', 'DIGITAL9', 'LIBRARY', 'TV_AV', 'VOLUMEUP',
    'VOLUMEDOWN', 'NETFLIX', 'YOUTUBE', 'PRIME_VIDEO', 'ACTION3', 'APPS',
    'FILES', 'MUTE', 'DISCOVERY', 'ASSERT', 'NOTASSERT',
])


# ─── 请求体模型 ────────────────────────────────────────────────────────────────

class ValidKeysUpdateRequest(BaseModel):
    keys: List[str]


class KeyCodesUpdateRequest(BaseModel):
    key_codes: Dict[str, int]


class CustomCommandsUpdateRequest(BaseModel):
    custom_commands: Dict[str, str]


class CreateSchemeRequest(BaseModel):
    name: str


class DuplicateSchemeRequest(BaseModel):
    new_name: str


# ─── 配置读写 ──────────────────────────────────────────────────────────────────

DEFAULT_EXTRA_COMMAND_DELAY = 0.0


def _normalize_extra_delay(value) -> float:
    """把任意输入归一化为非负 float；非法值回落到 0。"""
    try:
        delay = float(value)
    except (TypeError, ValueError):
        return DEFAULT_EXTRA_COMMAND_DELAY
    if delay < 0:
        return DEFAULT_EXTRA_COMMAND_DELAY
    return delay


DEFAULT_COLOR_MIN_SIMILARITY = 0.4
DEFAULT_COLOR_WEIGHT = 0.2
DEFAULT_FEATURE_MIN_SIMILARITY = 0.3


def _normalize_color_threshold(value) -> float:
    """把任意输入归一化为 0~1 的颜色相似度下限；非法值回落到默认 0.4。"""
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return DEFAULT_COLOR_MIN_SIMILARITY
    return parsed if 0.0 <= parsed <= 1.0 else DEFAULT_COLOR_MIN_SIMILARITY


def _normalize_color_weight(value) -> float:
    """把任意输入归一化为 0~1 的颜色权重；非法值回落到默认 0.2。"""
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return DEFAULT_COLOR_WEIGHT
    return parsed if 0.0 <= parsed <= 1.0 else DEFAULT_COLOR_WEIGHT


def _normalize_feature_threshold(value) -> float:
    """把任意输入归一化为 0~1 的特征相似度下限；非法值回落到默认 0.3。"""
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return DEFAULT_FEATURE_MIN_SIMILARITY
    return parsed if 0.0 <= parsed <= 1.0 else DEFAULT_FEATURE_MIN_SIMILARITY


def _normalize_config(data: dict) -> dict:
    schemes = data.get("schemes")
    if not isinstance(schemes, dict):
        schemes = {}

    normalized_schemes = {
        name: scheme if isinstance(scheme, dict) else {}
        for name, scheme in schemes.items()
        if isinstance(name, str) and name.strip()
    }

    if DEFAULT_SCHEME_NAME not in normalized_schemes:
        normalized_schemes[DEFAULT_SCHEME_NAME] = {}

    active_scheme = data.get("active_scheme")
    if not isinstance(active_scheme, str) or active_scheme not in normalized_schemes:
        active_scheme = DEFAULT_SCHEME_NAME

    return {
        "active_scheme": active_scheme,
        "schemes": normalized_schemes,
        # 全局：每条命令在用户指定 delay 之上再额外等待的秒数。0 表示不变。
        "extra_command_delay": _normalize_extra_delay(data.get("extra_command_delay")),
        # 全局：图片校验参数（颜色相似度下限 / 最终分颜色权重 / 特征相似度下限）。
        "color_min_similarity": _normalize_color_threshold(data.get("color_min_similarity")),
        "color_weight": _normalize_color_weight(data.get("color_weight")),
        "feature_min_similarity": _normalize_feature_threshold(data.get("feature_min_similarity")),
    }

def _load_config() -> dict:
    path = settings.CUSTOMIZATION_FILE
    if path.exists():
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            # 迁移：旧的扁平格式 → 新的方案格式
            if "schemes" not in data:
                scheme: dict = {}
                if isinstance(data.get("valid_keys"), list):
                    scheme["valid_keys"] = data["valid_keys"]
                if isinstance(data.get("key_codes"), dict):
                    scheme["key_codes"] = data["key_codes"]
                return _normalize_config({
                    "active_scheme": DEFAULT_SCHEME_NAME,
                    "schemes": {DEFAULT_SCHEME_NAME: scheme},
                })
            return _normalize_config(data)
        except Exception:
            pass
    return _normalize_config({})


def _save_config(data: dict) -> None:
    path = settings.CUSTOMIZATION_FILE
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def get_extra_command_delay() -> float:
    """供 ADB 命令执行链路读取的全局延迟增量，单位秒。

    每条命令的实际等待时间 = 用户给的 delay + 该值。例如用户设为 1，
    Excel 里写 ``HOME/1/0`` 实际会等 1 秒；写 ``HOME/1/2`` 实际会等 3 秒。
    """
    cfg = _load_config()
    return _normalize_extra_delay(cfg.get("extra_command_delay"))


def get_color_min_similarity() -> float:
    """图片校验的颜色相似度下限（0~1），供 image_service 运行时读取。"""
    cfg = _load_config()
    return _normalize_color_threshold(cfg.get("color_min_similarity"))


def get_color_weight() -> float:
    """图片校验最终分的颜色权重（0~1），供 image_service 运行时读取。"""
    cfg = _load_config()
    return _normalize_color_weight(cfg.get("color_weight"))


def get_feature_min_similarity() -> float:
    """图片校验的特征相似度下限（0~1），供 image_service 运行时读取。"""
    cfg = _load_config()
    return _normalize_feature_threshold(cfg.get("feature_min_similarity"))


# ─── 全局命令延迟增量 ──────────────────────────────────────────────────────────


class ExtraCommandDelayUpdateRequest(BaseModel):
    extra_command_delay: float


@router.get("/extra-command-delay")
def get_extra_command_delay_route():
    return {"extra_command_delay": get_extra_command_delay()}


@router.put("/extra-command-delay")
def update_extra_command_delay_route(req: ExtraCommandDelayUpdateRequest):
    delay = _normalize_extra_delay(req.extra_command_delay)
    config = _load_config()
    config["extra_command_delay"] = delay
    _save_config(config)
    return {"extra_command_delay": delay}


# ─── 图片校验颜色配置 ─────────────────────────────────────────────────────────

class ColorVerifyConfigUpdateRequest(BaseModel):
    color_min_similarity: Optional[float] = None
    color_weight: Optional[float] = None
    feature_min_similarity: Optional[float] = None


@router.get("/color-verify-config")
def get_color_verify_config_route():
    """获取图片校验的参数（颜色下限 / 颜色权重 / 特征下限）。"""
    return {
        "color_min_similarity": get_color_min_similarity(),
        "color_weight": get_color_weight(),
        "feature_min_similarity": get_feature_min_similarity(),
    }


@router.put("/color-verify-config")
def update_color_verify_config_route(req: ColorVerifyConfigUpdateRequest):
    """更新图片校验的参数（可只更新其中一个）。"""
    config = _load_config()
    if req.color_min_similarity is not None:
        config["color_min_similarity"] = _normalize_color_threshold(req.color_min_similarity)
    if req.color_weight is not None:
        config["color_weight"] = _normalize_color_weight(req.color_weight)
    if req.feature_min_similarity is not None:
        config["feature_min_similarity"] = _normalize_feature_threshold(req.feature_min_similarity)
    _save_config(config)
    return {
        "color_min_similarity": _normalize_color_threshold(config.get("color_min_similarity")),
        "color_weight": _normalize_color_weight(config.get("color_weight")),
        "feature_min_similarity": _normalize_feature_threshold(config.get("feature_min_similarity")),
    }


def _require_scheme(config: dict, name: str) -> dict:
    scheme = config.get("schemes", {}).get(name)
    if scheme is None:
        raise HTTPException(status_code=404, detail=f"方案 '{name}' 不存在")
    return scheme


# ─── 方案管理 ──────────────────────────────────────────────────────────────────

@router.get("/schemes")
def list_schemes():
    """列出所有方案及当前激活方案名"""
    config = _load_config()
    active = config.get("active_scheme", DEFAULT_SCHEME_NAME)
    schemes = config.get("schemes", {})
    return {
        "active_scheme": active,
        "schemes": [
            {
                "name": name,
                "is_active": name == active,
                "valid_keys_count": len(s.get("valid_keys") or DEFAULT_VALID_KEYS),
                "key_codes_count": len(s.get("key_codes", {})),
                "custom_commands_count": len(s.get("custom_commands", {})),
            }
            for name, s in schemes.items()
        ],
    }


@router.post("/schemes")
def create_scheme(req: CreateSchemeRequest):
    """新建方案（空方案，使用默认按键与键值）"""
    name = req.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="方案名称不能为空")
    config = _load_config()
    if name in config.get("schemes", {}):
        raise HTTPException(status_code=409, detail=f"方案 '{name}' 已存在")
    config.setdefault("schemes", {})[name] = {}
    if not config.get("active_scheme") or config["active_scheme"] not in config["schemes"]:
        config["active_scheme"] = name
    _save_config(config)
    return {"name": name, "active_scheme": config["active_scheme"]}


@router.delete("/schemes/{scheme_name}")
def delete_scheme(scheme_name: str):
    """删除方案（至少保留一个）"""
    config = _load_config()
    schemes = config.get("schemes", {})
    if scheme_name not in schemes:
        raise HTTPException(status_code=404, detail=f"方案 '{scheme_name}' 不存在")
    if len(schemes) <= 1:
        raise HTTPException(status_code=400, detail="至少需要保留一个方案")
    del schemes[scheme_name]
    if config.get("active_scheme") == scheme_name:
        config["active_scheme"] = next(iter(schemes))
    _save_config(config)
    return {"message": f"方案 '{scheme_name}' 已删除", "active_scheme": config["active_scheme"]}


@router.put("/schemes/{scheme_name}/activate")
def activate_scheme(scheme_name: str):
    """将指定方案设为激活方案（运行时读取该方案的配置）"""
    config = _load_config()
    _require_scheme(config, scheme_name)
    config["active_scheme"] = scheme_name
    _save_config(config)
    return {"active_scheme": scheme_name}


@router.post("/schemes/{scheme_name}/duplicate")
def duplicate_scheme(scheme_name: str, req: DuplicateSchemeRequest):
    """复制方案"""
    new_name = req.new_name.strip()
    if not new_name:
        raise HTTPException(status_code=400, detail="新方案名称不能为空")
    config = _load_config()
    source = _require_scheme(config, scheme_name)
    if new_name in config.get("schemes", {}):
        raise HTTPException(status_code=409, detail=f"方案 '{new_name}' 已存在")
    config["schemes"][new_name] = copy.deepcopy(source)
    _save_config(config)
    return {"name": new_name}


# ─── 合法按键名称 ──────────────────────────────────────────────────────────────

@router.get("/schemes/{scheme_name}/valid-keys")
def get_valid_keys(scheme_name: str):
    config = _load_config()
    scheme = _require_scheme(config, scheme_name)
    custom = scheme.get("valid_keys")
    if isinstance(custom, list) and custom:
        return {"keys": sorted(custom), "is_custom": True}
    return {"keys": DEFAULT_VALID_KEYS, "is_custom": False}


@router.put("/schemes/{scheme_name}/valid-keys")
def update_valid_keys(scheme_name: str, req: ValidKeysUpdateRequest):
    keys = sorted(set(k.strip().upper() for k in req.keys if k.strip()))
    if not keys:
        raise HTTPException(status_code=400, detail="按键列表不能为空")
    config = _load_config()
    _require_scheme(config, scheme_name)
    config["schemes"][scheme_name]["valid_keys"] = keys
    _save_config(config)
    return {"keys": keys}


@router.post("/schemes/{scheme_name}/valid-keys/reset")
def reset_valid_keys(scheme_name: str):
    config = _load_config()
    _require_scheme(config, scheme_name)
    config["schemes"][scheme_name].pop("valid_keys", None)
    _save_config(config)
    return {"keys": DEFAULT_VALID_KEYS}


# ─── 键值映射 ──────────────────────────────────────────────────────────────────

@router.get("/schemes/{scheme_name}/key-codes")
def get_key_codes(scheme_name: str):
    config = _load_config()
    scheme = _require_scheme(config, scheme_name)
    custom = {k.upper(): v for k, v in scheme.get("key_codes", {}).items()}
    merged = {**KEYCODE_MAP, **custom}
    return {
        "key_codes": dict(sorted(merged.items())),
        "custom_overrides": dict(sorted(custom.items())),
    }


@router.put("/schemes/{scheme_name}/key-codes")
def update_key_codes(scheme_name: str, req: KeyCodesUpdateRequest):
    validated: Dict[str, int] = {}
    for k, v in req.key_codes.items():
        key = k.strip().upper()
        if not key:
            raise HTTPException(status_code=400, detail="按键名称不能为空")
        if not isinstance(v, int) or v < 0:
            raise HTTPException(status_code=400, detail=f"'{key}' 的键值必须为非负整数")
        validated[key] = v
    config = _load_config()
    scheme = _require_scheme(config, scheme_name)
    scheme["key_codes"] = validated

    existing_valid_keys = scheme.get("valid_keys")
    if isinstance(existing_valid_keys, list) and existing_valid_keys:
        merged_valid_keys = {
            str(key).strip().upper()
            for key in existing_valid_keys
            if isinstance(key, str) and key.strip()
        }
    else:
        merged_valid_keys = set(DEFAULT_VALID_KEYS)
    merged_valid_keys.update(validated.keys())
    scheme["valid_keys"] = sorted(merged_valid_keys)

    _save_config(config)
    merged = {**KEYCODE_MAP, **validated}
    return {
        "key_codes": dict(sorted(merged.items())),
        "custom_overrides": dict(sorted(validated.items())),
    }


@router.delete("/schemes/{scheme_name}/key-codes/{key_name}")
def delete_key_code(scheme_name: str, key_name: str):
    config = _load_config()
    _require_scheme(config, scheme_name)
    key = key_name.strip().upper()
    overrides = config["schemes"][scheme_name].get("key_codes", {})
    if key not in overrides:
        raise HTTPException(status_code=404, detail=f"'{key}' 不是自定义键值")
    del overrides[key]
    config["schemes"][scheme_name]["key_codes"] = overrides
    _save_config(config)
    merged = {**KEYCODE_MAP, **overrides}
    return {
        "key_codes": dict(sorted(merged.items())),
        "custom_overrides": dict(sorted(overrides.items())),
    }


@router.post("/schemes/{scheme_name}/key-codes/reset")
def reset_key_codes(scheme_name: str):
    config = _load_config()
    _require_scheme(config, scheme_name)
    config["schemes"][scheme_name].pop("key_codes", None)
    _save_config(config)
    return {
        "key_codes": dict(sorted(KEYCODE_MAP.items())),
        "custom_overrides": {},
    }


@router.delete("/key-codes/{key_name}")
async def delete_key_code_override(key_name: str):
    """删除单个自定义键值覆盖（还原为默认值）"""
    key = key_name.strip().upper()
    cfg = _load_config()
    custom = cfg.get("key_codes", {})
    custom.pop(key, None)
    cfg["key_codes"] = custom
    _save_config(cfg)
    merged = {**KEYCODE_MAP, **{k.upper(): v for k, v in custom.items()}}
    return {
        "key_codes": dict(sorted(merged.items())),
        "custom_overrides": dict(sorted({k.upper(): v for k, v in custom.items()}.items())),
    }


@router.post("/key-codes/reset")
async def reset_key_codes():
    """清除所有自定义键值覆盖，还原为全部默认值"""
    cfg = _load_config()
    cfg.pop("key_codes", None)
    _save_config(cfg)
    return {
        "key_codes": dict(sorted(KEYCODE_MAP.items())),
        "custom_overrides": {},
    }


# ─── 自定义 ADB 命令 ──────────────────────────────────────────────────────────

def _normalize_custom_commands(commands: Dict[str, str]) -> Dict[str, str]:
    """校验并归一化自定义命令映射：键名大写、命令非空且不含换行。"""
    validated: Dict[str, str] = {}
    for name, command in (commands or {}).items():
        key = name.strip().upper()
        if not key:
            raise HTTPException(status_code=400, detail="按键名称不能为空")
        cmd = command.strip()
        if not cmd:
            raise HTTPException(status_code=400, detail=f"'{key}' 的命令不能为空")
        if "\n" in cmd or "\r" in cmd:
            raise HTTPException(status_code=400, detail=f"'{key}' 的命令不能包含换行符")
        validated[key] = cmd
    return validated


def _merge_commands_into_valid_keys(scheme: dict, command_keys: set) -> None:
    """把自定义命令按键名并入方案合法按键，避免 Excel 校验/回放时被当作无效按键。"""
    existing = scheme.get("valid_keys")
    if isinstance(existing, list) and existing:
        merged = {str(k).strip().upper() for k in existing if isinstance(k, str) and k.strip()}
    else:
        merged = set(DEFAULT_VALID_KEYS)
    merged.update(command_keys)
    scheme["valid_keys"] = sorted(merged)


@router.get("/schemes/{scheme_name}/custom-commands")
def get_custom_commands(scheme_name: str):
    """获取方案的按键名 → adb 命令映射。"""
    config = _load_config()
    scheme = _require_scheme(config, scheme_name)
    commands = {
        k.upper(): v
        for k, v in scheme.get("custom_commands", {}).items()
        if isinstance(v, str) and v.strip()
    }
    return {"custom_commands": dict(sorted(commands.items()))}


@router.put("/schemes/{scheme_name}/custom-commands")
def update_custom_commands(scheme_name: str, req: CustomCommandsUpdateRequest):
    """整体替换方案的按键名 → adb 命令映射，并把键名并入合法按键。"""
    validated = _normalize_custom_commands(req.custom_commands)
    config = _load_config()
    scheme = _require_scheme(config, scheme_name)
    scheme["custom_commands"] = validated
    _merge_commands_into_valid_keys(scheme, set(validated.keys()))
    _save_config(config)
    return {"custom_commands": dict(sorted(validated.items()))}


@router.delete("/schemes/{scheme_name}/custom-commands/{key_name}")
def delete_custom_command(scheme_name: str, key_name: str):
    """删除单条自定义命令。"""
    config = _load_config()
    scheme = _require_scheme(config, scheme_name)
    key = key_name.strip().upper()
    commands = scheme.get("custom_commands", {})
    if key not in commands:
        raise HTTPException(status_code=404, detail=f"'{key}' 不是自定义命令")
    del commands[key]
    scheme["custom_commands"] = commands
    _save_config(config)
    return {"custom_commands": dict(sorted(commands.items()))}


@router.post("/schemes/{scheme_name}/custom-commands/reset")
def reset_custom_commands(scheme_name: str):
    """清空方案的自定义命令。"""
    config = _load_config()
    scheme = _require_scheme(config, scheme_name)
    scheme.pop("custom_commands", None)
    _save_config(config)
    return {"custom_commands": {}}


# ─── sendevent 长按配置 ───────────────────────────────────────────────────────

LONG_PRESS_METHODS = ("auto", "sendevent", "input")


def _normalize_long_press_method(value) -> str:
    method = str(value or "auto").strip().lower()
    return method if method in LONG_PRESS_METHODS else "auto"


class SendeventConfigUpdateRequest(BaseModel):
    sendevent_device: Optional[str] = None
    sendevent_keycode_overrides: Optional[Dict[str, int]] = None
    long_press_method: Optional[str] = None


@router.get("/schemes/{scheme_name}/sendevent-config")
def get_sendevent_config(scheme_name: str):
    """获取方案的 sendevent 长按配置。"""
    config = _load_config()
    scheme = _require_scheme(config, scheme_name)
    return {
        "sendevent_device": scheme.get("sendevent_device", ""),
        "sendevent_keycode_overrides": scheme.get("sendevent_keycode_overrides", {}),
        "long_press_method": _normalize_long_press_method(scheme.get("long_press_method")),
    }


@router.put("/schemes/{scheme_name}/sendevent-config")
def update_sendevent_config(scheme_name: str, req: SendeventConfigUpdateRequest):
    """更新方案的 sendevent 长按配置。"""
    config = _load_config()
    scheme = _require_scheme(config, scheme_name)

    if req.sendevent_device is not None:
        device = req.sendevent_device.strip()
        if device:
            scheme["sendevent_device"] = device
        else:
            scheme.pop("sendevent_device", None)

    if req.sendevent_keycode_overrides is not None:
        validated: Dict[str, int] = {}
        for k, v in req.sendevent_keycode_overrides.items():
            if not isinstance(v, int) or v < 0:
                raise HTTPException(status_code=400, detail=f"Linux keycode 必须为非负整数，收到: {v}")
            validated[str(k)] = v
        scheme["sendevent_keycode_overrides"] = validated

    if req.long_press_method is not None:
        method = _normalize_long_press_method(req.long_press_method)
        if method != "auto":
            scheme["long_press_method"] = method
        else:
            scheme.pop("long_press_method", None)

    _save_config(config)
    return {
        "sendevent_device": scheme.get("sendevent_device", ""),
        "sendevent_keycode_overrides": scheme.get("sendevent_keycode_overrides", {}),
        "long_press_method": _normalize_long_press_method(scheme.get("long_press_method")),
    }


@router.post("/schemes/{scheme_name}/sendevent-config/reset")
def reset_sendevent_config(scheme_name: str):
    """重置方案的 sendevent 配置为默认值。"""
    config = _load_config()
    _require_scheme(config, scheme_name)
    config["schemes"][scheme_name].pop("sendevent_device", None)
    config["schemes"][scheme_name].pop("sendevent_keycode_overrides", None)
    config["schemes"][scheme_name].pop("long_press_method", None)
    _save_config(config)
    return {"sendevent_device": "", "sendevent_keycode_overrides": {}, "long_press_method": "auto"}


# ─── Excel 导入：从 RC 表 RC_FUN/KeyCode 列生成新方案 ───────────────────────────

EXCEL_IMPORT_TARGET_SHEET = "RC"
EXCEL_IMPORT_KEY_COLUMN = "RC_FUN"
EXCEL_IMPORT_VALUE_COLUMN = "KeyCode"
EXCEL_IMPORT_MAX_NAME_LENGTH = 30


def _normalize_sheet_name(name: str) -> str:
    return str(name or "").strip().lower()


def _normalize_header_label(value) -> str:
    return str(value or "").strip().lower()


def _coerce_keycode(value) -> Optional[int]:
    """把 Excel 单元格里的 KeyCode 转成非负整数；解析失败返回 None。"""
    if value is None:
        return None
    if isinstance(value, bool):
        return None  # ``True``/``False`` 被 openpyxl 当 1/0，但语义不对
    if isinstance(value, int):
        return value if value >= 0 else None
    if isinstance(value, float):
        if value.is_integer() and value >= 0:
            return int(value)
        return None
    text = str(value).strip()
    if not text:
        return None
    # 0x... 形式（不少厂商表格里会用十六进制）
    try:
        if text.lower().startswith(("0x", "&h", "h")):
            stripped = text.lstrip("hH").lstrip("&").lstrip("0xX")
            return int(stripped, 16) if stripped else None
        return int(text, 10)
    except ValueError:
        return None


def _resolve_rc_sheet(workbook):
    """优先精确匹配 'RC'，其次大小写不敏感匹配。"""
    if EXCEL_IMPORT_TARGET_SHEET in workbook.sheetnames:
        return workbook[EXCEL_IMPORT_TARGET_SHEET]
    target = _normalize_sheet_name(EXCEL_IMPORT_TARGET_SHEET)
    for sheet_name in workbook.sheetnames:
        if _normalize_sheet_name(sheet_name) == target:
            return workbook[sheet_name]
    return None


def _resolve_rc_columns(worksheet) -> Tuple[Optional[int], Optional[int]]:
    """返回 (RC_FUN 列索引, KeyCode 列索引)；找不到返回 None。"""
    if worksheet.max_row is None or worksheet.max_row < 1:
        return None, None
    key_target = _normalize_header_label(EXCEL_IMPORT_KEY_COLUMN)
    value_target = _normalize_header_label(EXCEL_IMPORT_VALUE_COLUMN)
    key_col, value_col = None, None
    max_col = worksheet.max_column or 0
    for col_index in range(1, max_col + 1):
        cell_value = worksheet.cell(row=1, column=col_index).value
        label = _normalize_header_label(cell_value)
        if not label:
            continue
        if key_col is None and label == key_target:
            key_col = col_index
        elif value_col is None and label == value_target:
            value_col = col_index
        if key_col and value_col:
            break
    return key_col, value_col


def _parse_rc_excel(file_bytes: bytes) -> Dict[str, int]:
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"无法解析 Excel 文件: {exc}") from exc

    try:
        worksheet = _resolve_rc_sheet(workbook)
        if worksheet is None:
            raise HTTPException(
                status_code=400,
                detail=f"Excel 文件中没有名为 '{EXCEL_IMPORT_TARGET_SHEET}' 的工作表",
            )

        key_col, value_col = _resolve_rc_columns(worksheet)
        missing = []
        if key_col is None:
            missing.append(EXCEL_IMPORT_KEY_COLUMN)
        if value_col is None:
            missing.append(EXCEL_IMPORT_VALUE_COLUMN)
        if missing:
            raise HTTPException(
                status_code=400,
                detail=f"RC 表中缺少必要的列: {', '.join(missing)}",
            )

        key_codes: Dict[str, int] = {}
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            raw_key = row[key_col - 1] if key_col - 1 < len(row) else None
            raw_value = row[value_col - 1] if value_col - 1 < len(row) else None
            normalized_key = str(raw_key or "").strip().upper()
            if not normalized_key:
                continue
            keycode = _coerce_keycode(raw_value)
            if keycode is None:
                continue
            # 同 key 多次出现时保留第一条；用户表里如果有冲突先尊重靠前的那一行
            key_codes.setdefault(normalized_key, keycode)

        if not key_codes:
            raise HTTPException(
                status_code=400,
                detail="RC 表中没有解析到任何有效的 RC_FUN/KeyCode 配对",
            )
        return key_codes
    finally:
        try:
            workbook.close()
        except Exception:
            pass


def _sanitize_scheme_name(raw: str) -> str:
    name = str(raw or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="方案名称不能为空")
    if len(name) > EXCEL_IMPORT_MAX_NAME_LENGTH:
        raise HTTPException(
            status_code=400,
            detail=f"方案名称长度不能超过 {EXCEL_IMPORT_MAX_NAME_LENGTH} 个字符",
        )
    if re.search(r"[\\/:*?\"<>|]", name):
        raise HTTPException(
            status_code=400,
            detail="方案名称不能包含 \\ / : * ? \" < > |",
        )
    return name


def _resolve_unique_scheme_name(config: dict, base: str) -> str:
    schemes = config.get("schemes", {})
    if base not in schemes:
        return base
    index = 1
    while True:
        candidate = f"{base} (导入{index})"
        if len(candidate) > EXCEL_IMPORT_MAX_NAME_LENGTH:
            stem_max = EXCEL_IMPORT_MAX_NAME_LENGTH - len(f" (导入{index})")
            candidate = f"{base[:max(1, stem_max)]} (导入{index})"
        if candidate not in schemes:
            return candidate
        index += 1


@router.post("/schemes/import-excel")
async def import_scheme_from_excel(
    file: UploadFile = File(...),
    scheme_name: Optional[str] = Form(default=None),
    conflict: str = Form("rename"),
):
    """从 Excel RC 表导入键值映射，生成新方案。

    - 工作表必须叫 ``RC``
    - 必须有 ``RC_FUN`` 与 ``KeyCode`` 两列（首行作为表头）
    - 方案名默认取文件名（去扩展名），可通过 ``scheme_name`` 覆盖
    - 冲突策略 ``conflict``：``rename`` / ``overwrite`` / ``skip``
    """
    if conflict not in {"rename", "overwrite", "skip"}:
        raise HTTPException(status_code=400, detail=f"未知的冲突处理策略: {conflict}")

    try:
        raw_bytes = await file.read()
    finally:
        await file.close()

    if not raw_bytes:
        raise HTTPException(status_code=400, detail="导入文件为空")

    upload_filename = file.filename or ""
    if not upload_filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx/.xlsm 格式的 Excel 文件")

    key_codes = _parse_rc_excel(raw_bytes)

    # 默认方案名 = 文件名 (去扩展名)；用户可主动覆盖
    fallback_stem = re.sub(r"\.[^.]+$", "", upload_filename).strip() or "RC"
    requested_name = _sanitize_scheme_name(scheme_name or fallback_stem)

    config = _load_config()
    final_name = requested_name
    renamed_from: Optional[str] = None
    skipped = False

    if requested_name in config.get("schemes", {}):
        if conflict == "overwrite":
            final_name = requested_name
        elif conflict == "skip":
            skipped = True
        else:  # rename
            final_name = _resolve_unique_scheme_name(config, requested_name)
            if final_name != requested_name:
                renamed_from = requested_name

    if not skipped:
        # 同时把这些按键加入 valid_keys，避免下次回放因"按键无效"被跳过
        merged_valid_keys = set(DEFAULT_VALID_KEYS)
        merged_valid_keys.update(key_codes.keys())
        config.setdefault("schemes", {})[final_name] = {
            "key_codes": dict(sorted(key_codes.items())),
            "valid_keys": sorted(merged_valid_keys),
        }
        if not config.get("active_scheme") or config["active_scheme"] not in config["schemes"]:
            config["active_scheme"] = final_name
        _save_config(config)

    schemes = config.get("schemes", {})
    return {
        "status": "success",
        "imported_scheme": None if skipped else final_name,
        "renamed_from": renamed_from,
        "skipped": skipped,
        "key_codes_count": len(key_codes),
        "active_scheme": config.get("active_scheme"),
        "schemes": [
            {
                "name": name,
                "is_active": name == config.get("active_scheme"),
                "valid_keys_count": len(s.get("valid_keys") or DEFAULT_VALID_KEYS),
                "key_codes_count": len(s.get("key_codes", {})),
            }
            for name, s in sorted(schemes.items())
        ],
    }
