"""Excel服务模块"""
import re
from copy import copy
from datetime import datetime
from pathlib import Path
import pandas as pd
from typing import Dict, Any, List, Optional
from openpyxl import load_workbook
from ..config import settings
from ..utils.adb_controller import ADBController
from ..utils.validators import ExcelValidator
from ..utils.path_resolver import list_excel_files, resolve_excel_file

class ExcelService:
    """Excel服务类"""

    OPENPYXL_SUPPORTED_SUFFIXES = {'.xlsx', '.xlsm', '.xltx', '.xltm'}

    def __init__(self):
        self.controller = ADBController()
        self.validator = ExcelValidator()

    def get_excel_files(self, directory: str = None) -> List[str]:
        """获取目录下的Excel文件列表"""
        if directory is None:
            return list_excel_files()

        target_dir = Path(directory)
        if not target_dir.exists() or not target_dir.is_dir():
            return []

        excel_files = []
        for file_path in target_dir.iterdir():
            if file_path.is_file() and file_path.suffix.lower() in {'.xlsx', '.xls'}:
                excel_files.append(file_path.name)
        return sorted(excel_files)

    def validate(self, file_name: str) -> Dict[str, Any]:
        """验证Excel文件"""
        file_path = resolve_excel_file(file_name)
        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_name}")

        return self.validator.validate(str(file_path))

    def analyze(self, file_name: str) -> Dict[str, Any]:
        """分析Excel文件"""
        file_path = resolve_excel_file(file_name)
        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_name}")

        return self.controller.read_excel_commands(str(file_path))

    def get_case_state(self, file_name: str, case_number: str) -> Dict[str, Any]:
        """查询指定 case 在 Excel 中的当前状态。

        返回该 case 是否已存在、已有多少次 Assert、checkPic 数量等信息，
        供前端在写入前判断是否需要弹出 Assert 格式确认框。
        """
        normalized_case = self._normalize_case_number(case_number)
        file_path = resolve_excel_file(file_name)
        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_name}")

        if not normalized_case:
            return {"exists": False, "assert_count": 0, "check_pic_count": 0, "prescript": "", "check_pic": ""}

        try:
            if not self._supports_cell_level_write(file_path):
                return self._get_case_state_with_pandas(file_path, normalized_case)
            return self._get_case_state_with_openpyxl(file_path, normalized_case)
        except Exception as e:
            raise Exception(f"查询 case 状态失败: {e}")

    def _get_case_state_with_pandas(self, file_path: Path, normalized_case: str) -> Dict[str, Any]:
        """pandas 版：读取指定 case 的 preScript 和 checkPic 状态。"""
        df = pd.read_excel(file_path)
        row_index = self._find_existing_row_for_case_with_pandas(df, normalized_case)
        if row_index is None:
            return {"exists": False, "assert_count": 0, "check_pic_count": 0, "prescript": "", "check_pic": ""}

        prescript = str(df.loc[row_index, 'preScript'] or '') if 'preScript' in df.columns else ''
        check_pic_col = 'checkPic' if 'checkPic' in df.columns else 'verify_image'
        check_pic = str(df.loc[row_index, check_pic_col] or '') if check_pic_col in df.columns else ''

        return {
            "exists": True,
            "assert_count": prescript.upper().count('ASSERT/'),
            "check_pic_count": self._count_existing_check_pic_images(check_pic),
            "prescript": prescript,
            "check_pic": check_pic,
        }

    def _get_case_state_with_openpyxl(self, file_path: Path, normalized_case: str) -> Dict[str, Any]:
        """openpyxl 版：读取指定 case 的 preScript 和 checkPic 状态。"""
        workbook, worksheet = self._load_workbook_sheet(file_path)
        try:
            header_map = self._get_header_map(worksheet)
            excel_row = self._find_existing_row_for_case_in_sheet(worksheet, header_map, normalized_case)
            if excel_row is None:
                return {"exists": False, "assert_count": 0, "check_pic_count": 0, "prescript": "", "check_pic": ""}

            prescript_col = header_map.get('preScript')
            prescript = str(worksheet.cell(row=excel_row, column=prescript_col).value or '') if prescript_col else ''

            check_pic_col_name = 'checkPic' if 'checkPic' in header_map else 'verify_image'
            check_pic_col = header_map.get(check_pic_col_name)
            check_pic = str(worksheet.cell(row=excel_row, column=check_pic_col).value or '') if check_pic_col else ''

            return {
                "exists": True,
                "assert_count": prescript.upper().count('ASSERT/'),
                "check_pic_count": self._count_existing_check_pic_images(check_pic),
                "prescript": prescript,
                "check_pic": check_pic,
            }
        finally:
            workbook.close()

    def append_assert_to_case(self, file_name: str, case_number: str, assert_format: str = 'Assert/1/1') -> Dict[str, Any]:
        """直接将 Assert 指令写入指定 case 的 preScript 末尾。

        如果 preScript 末尾已经是 Assert 指令，则覆盖它；否则追加。
        """
        normalized_case = self._normalize_case_number(case_number)
        file_path = resolve_excel_file(file_name)
        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_name}")
        if not normalized_case:
            raise ValueError('用例编号不能为空')

        resolved_assert = (assert_format or '').strip() or 'Assert/1/1'

        try:
            if not self._supports_cell_level_write(file_path):
                return self._append_assert_with_pandas(file_path, normalized_case, resolved_assert)
            return self._append_assert_with_openpyxl(file_path, normalized_case, resolved_assert)
        except Exception as e:
            raise Exception(f"写入 Assert 失败: {e}")

    @staticmethod
    def _strip_trailing_assert(prescript: str) -> str:
        """移除 preScript 末尾的 Assert 指令（如有）。"""
        parts = [p.strip() for p in str(prescript or '').split(',') if p.strip()]
        if parts and re.match(r'^ASSERT/\d+/', parts[-1], re.IGNORECASE):
            parts.pop()
        return ','.join(parts)

    def _append_assert_with_pandas(self, file_path: Path, normalized_case: str, assert_format: str) -> Dict[str, Any]:
        """pandas 版：向已有 case 的 preScript 末尾写入 Assert。"""
        df = pd.read_excel(file_path)
        row_index = self._find_existing_row_for_case_with_pandas(df, normalized_case)
        if row_index is None:
            raise ValueError(f'未找到用例 {normalized_case} 的已写入行')

        if 'preScript' not in df.columns:
            df['preScript'] = None
        df['preScript'] = df['preScript'].astype(object)

        old_prescript = str(df.loc[row_index, 'preScript'] or '')
        stripped = self._strip_trailing_assert(old_prescript)
        new_prescript = f'{stripped},{assert_format}' if stripped else assert_format
        df.loc[row_index, 'preScript'] = new_prescript

        df.to_excel(file_path, index=False)
        return {
            "status": "ok",
            "file": file_path.name,
            "case_number": normalized_case,
            "prescript": new_prescript,
            "excel_row": row_index + 2,
            "data_row_index": row_index,
            "replaced": stripped != old_prescript.strip(),
        }

    def _append_assert_with_openpyxl(self, file_path: Path, normalized_case: str, assert_format: str) -> Dict[str, Any]:
        """openpyxl 版：向已有 case 的 preScript 末尾写入 Assert。"""
        workbook, worksheet = self._load_workbook_sheet(file_path)
        try:
            header_map = self._get_header_map(worksheet)
            excel_row = self._find_existing_row_for_case_in_sheet(worksheet, header_map, normalized_case)
            if excel_row is None:
                raise ValueError(f'未找到用例 {normalized_case} 的已写入行')

            pre_script_col = header_map.get('preScript')
            if not pre_script_col:
                raise ValueError('Excel 中没有 preScript 列')

            old_prescript = str(worksheet.cell(row=excel_row, column=pre_script_col).value or '')
            stripped = self._strip_trailing_assert(old_prescript)
            new_prescript = f'{stripped},{assert_format}' if stripped else assert_format
            self._write_sheet_cell(worksheet, excel_row, pre_script_col, new_prescript)

            workbook.save(file_path)
            return {
                "status": "ok",
                "file": file_path.name,
                "case_number": normalized_case,
                "prescript": new_prescript,
                "excel_row": excel_row,
                "data_row_index": excel_row - 2,
                "replaced": stripped != old_prescript.strip(),
            }
        finally:
            workbook.close()

    def preview(self, file_name: str, limit: int = 100) -> Dict[str, Any]:
        """预览Excel文件内容，返回列名和前几行数据。"""
        file_path = resolve_excel_file(file_name)
        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_name}")

        try:
            df = pd.read_excel(file_path)
            preview_df = df.head(limit).copy()
            preview_df = preview_df.where(pd.notna(preview_df), None)

            return {
                "columns": [str(col) for col in preview_df.columns.tolist()],
                "rows": preview_df.to_dict(orient="records"),
                "row_count": int(len(df))
            }
        except Exception as e:
            raise Exception(f"预览失败: {e}")

    def read_commands(self, file_name: str, row_index: int) -> Dict[str, Any]:
        """读取指定行的命令"""
        file_path = resolve_excel_file(file_name)
        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_name}")

        return self.controller.read_excel_commands(str(file_path), row_index)

    @staticmethod
    def _load_workbook_sheet(file_path: Path):
        workbook = load_workbook(
            file_path,
            keep_vba=file_path.suffix.lower() in {'.xlsm', '.xltm'},
        )
        worksheet = workbook.worksheets[0] if workbook.worksheets else workbook.active
        return workbook, worksheet

    @classmethod
    def _supports_cell_level_write(cls, file_path: Path) -> bool:
        return file_path.suffix.lower() in cls.OPENPYXL_SUPPORTED_SUFFIXES

    @staticmethod
    def _get_header_map(worksheet) -> Dict[str, int]:
        header_map = {}
        for column_index in range(1, worksheet.max_column + 1):
            header_value = worksheet.cell(row=1, column=column_index).value
            if header_value is None:
                continue

            header_name = str(header_value).strip()
            if header_name and header_name not in header_map:
                header_map[header_name] = column_index

        return header_map

    @staticmethod
    def _clone_cell_style(source_cell, target_cell) -> None:
        if source_cell is None or not source_cell.has_style:
            return

        target_cell._style = copy(source_cell._style)

    def _ensure_column(self, worksheet, header_map: Dict[str, int], column_name: str) -> int:
        existing_index = header_map.get(column_name)
        if existing_index is not None:
            return existing_index

        column_index = worksheet.max_column + 1 if worksheet.max_column else 1
        header_cell = worksheet.cell(row=1, column=column_index)
        if column_index > 1:
            self._clone_cell_style(worksheet.cell(row=1, column=column_index - 1), header_cell)
        header_cell.value = column_name
        header_map[column_name] = column_index
        return column_index

    def _prepare_target_cell(self, worksheet, row_index: int, column_index: int):
        target_cell = worksheet.cell(row=row_index, column=column_index)
        if target_cell.has_style:
            return target_cell

        style_source = None

        if column_index > 1:
            left_cell = worksheet.cell(row=row_index, column=column_index - 1)
            if left_cell.has_style:
                style_source = left_cell

        if style_source is None and row_index > 2:
            above_cell = worksheet.cell(row=row_index - 1, column=column_index)
            if above_cell.has_style:
                style_source = above_cell

        if style_source is not None:
            self._clone_cell_style(style_source, target_cell)

        return target_cell

    def _write_sheet_cell(self, worksheet, row_index: int, column_index: int, value: Any) -> None:
        target_cell = self._prepare_target_cell(worksheet, row_index, column_index)
        target_cell.value = value

    @staticmethod
    def _has_meaningful_cell_value(value: Any) -> bool:
        if value is None:
            return False

        if pd.isna(value):
            return False

        normalized = str(value).strip()
        return normalized != '' and normalized.lower() != 'nan'

    @staticmethod
    def _normalize_command_sequence(sequence: Any) -> str:
        return ','.join(
            part.strip()
            for part in str(sequence or '').split(',')
            if part.strip()
        )

    @staticmethod
    def _normalize_case_number(case_number: Any) -> str:
        normalized = str(case_number or '').strip()
        if not normalized or normalized.lower() == 'nan':
            return ''
        return normalized

    @staticmethod
    def _sanitize_file_stem(value: str) -> str:
        sanitized = ''.join('_' if char in '\\/:*?"<>|' else char for char in str(value or '').strip())
        sanitized = sanitized.strip().strip('.')
        return sanitized or 'screenshot'

    def _build_unique_check_pic_name(self, case_number: str) -> str:
        """第一次为某 case 写入时使用：直接返回 ``<case_number>.png``，不再做磁盘冲突避让。

        磁盘上是否已有同名旧文件不影响这里的命名（同 case 的截图本就该共用同一个 base
        名）。同 case 后续多次写入会走 ``_build_next_check_pic_name``，按 -1 / -2 递增。
        """
        normalized_case_number = self._normalize_case_number(case_number)
        if normalized_case_number:
            base_stem = self._sanitize_file_stem(normalized_case_number)
        else:
            base_stem = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f'{base_stem}.png'

    def _build_next_check_pic_name(self, existing_check_pic: str, case_number: str) -> str:
        """对同一 case 二次写入：基于 Excel 里 checkPic 单元格的"最后一个"图片名末尾的 -N，
        生成 -N+1 版本。``existing_check_pic`` 可能是逗号分隔的多张图（之前累加过的），
        这里取最后一张作为参照。

        必须以 ``case_number`` 作为已知前缀来识别 -N，否则像 ``CASE-42.png`` 这种 case
        本身就以数字结尾的会被错误地把 ``-42`` 当成后缀去递增。

        - existing="OSF10_ADV_0002.png", case="OSF10_ADV_0002" → "OSF10_ADV_0002-1.png"
        - existing="OSF10_ADV_0002.png,OSF10_ADV_0002-1.png", case="OSF10_ADV_0002"
            → "OSF10_ADV_0002-2.png"
        - existing="CASE-42.png", case="CASE-42" → "CASE-42-1.png"  # 不会误把 -42 当后缀
        - existing 为空 → "<case>-1.png"
        - 最后一张与 case_stem 对不上 → 退化到 "<case>-1.png"
        """
        case_stem = self._sanitize_file_stem(self._normalize_case_number(case_number) or 'screenshot')

        existing_raw = str(existing_check_pic or '').strip()
        if not existing_raw:
            return f'{case_stem}-1.png'

        # checkPic 可能已经是逗号分隔的多张图，取最后一张作为参照
        last_part = existing_raw.split(',')[-1].strip()
        if not last_part:
            return f'{case_stem}-1.png'

        path = Path(last_part)
        stem = path.stem or case_stem
        suffix = path.suffix or '.png'

        if stem == case_stem:
            return f'{case_stem}-1{suffix}'

        prefix = case_stem + '-'
        if stem.startswith(prefix):
            tail = stem[len(prefix):]
            if tail.isdigit():
                return f'{case_stem}-{int(tail) + 1}{suffix}'

        # 旧值与当前 case 对不上（用户改过 case_number 或外部修改），保守从 -1 开始
        return f'{case_stem}-1{suffix}'

    def _append_check_pic_value(self, existing_check_pic: str, next_check_pic: str) -> str:
        """把新图片名拼接到旧 checkPic 单元格末尾，逗号分隔；空值跳过。

        旧值里如果有段被早期版本错误地写入了绝对路径（如 ``D:\\Checkphoto\\xx.png``），
        这里会自动 normalize 成 basename，避免污染继续传递。
        """
        new_name = Path(str(next_check_pic or '').strip()).name if str(next_check_pic or '').strip() else ''

        existing_raw = str(existing_check_pic or '').strip().rstrip(',').strip()
        cleaned_parts: list[str] = []
        if existing_raw:
            for piece in existing_raw.split(','):
                piece = piece.strip()
                if not piece:
                    continue
                cleaned_parts.append(Path(piece).name)
        existing_normalized = ','.join(cleaned_parts)

        if not new_name:
            return existing_normalized
        if not existing_normalized:
            return new_name
        return f'{existing_normalized},{new_name}'

    def _build_extended_prescript(self, existing_prescript: str, new_sequence_compressed: str, assert_format: str = 'Assert/1/1') -> str:
        """对同一 case 二次写入：在旧 preScript 末尾追加新序列和 Assert。

        顺序：旧 preScript → 新序列 → Assert（如果用户输入了）。
        旧 preScript 已经包含之前的 Assert，所以只需要在新序列后面添加。

        Args:
            assert_format: 用户自定义的 assert 格式，默认 'Assert/1/1'。如果为空字符串则不插入 Assert。
        """
        pieces: list[str] = []
        existing_normalized = self._normalize_command_sequence(existing_prescript)
        if existing_normalized:
            pieces.append(existing_normalized.strip(','))
        new_clean = (new_sequence_compressed or '').strip(',')
        if new_clean:
            pieces.append(new_clean)
        # 只有当 assert_format 非空时才在最后插入 Assert
        assert_to_use = assert_format.strip()
        if assert_to_use:
            pieces.append(assert_to_use)
        return ','.join(p for p in pieces if p)

    @staticmethod
    def _parse_assert_count(assert_format: str) -> int:
        """从 assert 格式字符串中解析 count 值。

        ``Assert/2/1`` → 2，``Assert/3/1`` → 3，解析失败返回 0。
        """
        match = re.match(r'Assert/(\d+)/', assert_format or '')
        return int(match.group(1)) if match else 0

    def _parse_total_assert_count(self, prescript: str) -> int:
        """解析 preScript 中所有 Assert 的 x 值之和。

        例如 ``DOWN/1/7,Assert/1/1,Youtube/1/1,Assert/2/1`` 返回 3（1+2）。
        """
        if not prescript:
            return 0
        total = 0
        for part in str(prescript).split(','):
            part = part.strip()
            match = re.match(r'Assert/(\d+)/', part)
            if match:
                total += int(match.group(1))
        return total

    def _build_check_pic_list_for_assert_count(self, case_number: str, assert_count: int, start_index: int = 0) -> List[str]:
        """根据 assert 的 count 值和起始索引生成 checkPic 文件名列表。

        ``Assert/N/1`` 的 N 代表本次写入新增 N 张图片，编号从已有图片数之后继续：
        - start_index == 0, count == 1 → ``[<case>.png]``
        - start_index == 1, count == 1 → ``[<case>-1.png]``
        - start_index == 1, count == 2 → ``[<case>-1.png, <case>-2.png]``
        - start_index == 0, count == 3 → ``[<case>.png, <case>-1.png, <case>-2.png]``
        """
        case_stem = self._sanitize_file_stem(self._normalize_case_number(case_number) or 'screenshot')
        count = assert_count if assert_count >= 1 else 1
        result: List[str] = []
        for i in range(count):
            index = start_index + i
            if index == 0:
                result.append(f'{case_stem}.png')
            else:
                result.append(f'{case_stem}-{index}.png')
        return result

    @staticmethod
    def _count_existing_check_pic_images(existing_check_pic: str) -> int:
        """计算 checkPic 单元格中已有图片数量。"""
        raw = str(existing_check_pic or '').strip()
        if not raw:
            return 0
        return len([p for p in raw.split(',') if p.strip()])

    def _build_check_pic_for_append(self, existing_check_pic: str, case_number: str, assert_count: int) -> str:
        """追加写入时生成新的 checkPic 值。

        逻辑：已有 K 张图片，本次 assert 新增 N 张。
        从第 K 张之后开始编号（K, K+1, …, K+N-1），避免重复已有图片。
        """
        existing_count = self._count_existing_check_pic_images(existing_check_pic)
        new_images = self._build_check_pic_list_for_assert_count(case_number, assert_count, start_index=existing_count)
        return self._append_check_pic_list(existing_check_pic, new_images)

    @staticmethod
    def _append_check_pic_list(existing_check_pic: str, new_images: List[str]) -> str:
        """把多张图片名逐个拼接到 checkPic 单元格末尾，逗号分隔。"""
        existing_raw = str(existing_check_pic or '').strip().rstrip(',').strip()
        cleaned: List[str] = []
        if existing_raw:
            for piece in existing_raw.split(','):
                piece = piece.strip()
                if piece:
                    cleaned.append(Path(piece).name)
        for img in new_images:
            name = Path(str(img or '').strip()).name if str(img or '').strip() else ''
            if name:
                cleaned.append(name)
        return ','.join(cleaned)

    def _candidate_case_id_columns(self, available_columns) -> list[str]:
        """匹配 case_number 时按优先级使用的列名。"""
        ordered = []
        for column_name in ('testID', 'title', 'testItem'):
            if column_name in available_columns and column_name not in ordered:
                ordered.append(column_name)
        return ordered

    def _find_existing_row_for_case_with_pandas(
        self,
        df: pd.DataFrame,
        normalized_case_number: str,
    ) -> Optional[int]:
        """从下往上找第一行：testID/title 匹配 case_number 且 preScript 已有内容。"""
        if not normalized_case_number:
            return None
        case_id_columns = self._candidate_case_id_columns(df.columns)
        if not case_id_columns:
            return None
        if 'preScript' not in df.columns:
            return None

        target = str(normalized_case_number).strip()
        for data_row_index in range(len(df) - 1, -1, -1):
            row = df.iloc[data_row_index]
            matched = any(
                str(row.get(column_name) or '').strip() == target
                for column_name in case_id_columns
            )
            if not matched:
                continue
            if self._has_meaningful_cell_value(row.get('preScript')):
                return data_row_index
        return None

    def _find_existing_row_for_case_in_sheet(
        self,
        worksheet,
        header_map: Dict[str, int],
        normalized_case_number: str,
    ) -> Optional[int]:
        """openpyxl 版：从最后一行往上找 case 匹配且 preScript 非空的行号（excel_row）。"""
        if not normalized_case_number:
            return None
        case_id_columns = self._candidate_case_id_columns(header_map.keys())
        if not case_id_columns:
            return None
        if 'preScript' not in header_map:
            return None

        pre_script_column_index = header_map['preScript']
        target = str(normalized_case_number).strip()
        for excel_row in range(worksheet.max_row, 1, -1):
            matched = False
            for column_name in case_id_columns:
                column_index = header_map.get(column_name)
                if not column_index:
                    continue
                value = worksheet.cell(row=excel_row, column=column_index).value
                if str(value or '').strip() == target:
                    matched = True
                    break
            if not matched:
                continue
            pre_script_value = worksheet.cell(row=excel_row, column=pre_script_column_index).value
            if self._has_meaningful_cell_value(pre_script_value):
                return excel_row
        return None

    @classmethod
    def _compress_adjacent_command_sequence(cls, sequence: Any) -> str:
        normalized_sequence = cls._normalize_command_sequence(sequence)
        if not normalized_sequence:
            return ''

        parts = [part.strip() for part in normalized_sequence.split(',') if part.strip()]
        compressed_parts: List[str] = []

        for part in parts:
            segments = part.split('/')
            if len(segments) < 3:
                compressed_parts.append(part)
                continue

            key = segments[0]
            try:
                count = int(segments[1])
            except (TypeError, ValueError):
                count = 1
            if count <= 0:
                count = 1

            delay = segments[2]
            if delay == '*':
                compressed_parts.append(f"{key}/{count}/{delay}")
                continue

            last = compressed_parts[-1] if compressed_parts else None
            if last:
                last_segments = last.split('/')
                if len(last_segments) >= 3 and last_segments[0] == key and last_segments[2] == delay:
                    try:
                        last_count = int(last_segments[1])
                    except (TypeError, ValueError):
                        last_count = 1
                    if last_count <= 0:
                        last_count = 1

                    compressed_parts[-1] = f"{key}/{last_count + count}/{delay}"
                    continue

            compressed_parts.append(f"{key}/{count}/{delay}")

        return ','.join(compressed_parts)

    def _find_latest_blank_prescript_data_row_with_pandas(self, df: pd.DataFrame) -> int:
        if 'preScript' not in df.columns:
            df['preScript'] = None

        for data_row_index in range(len(df) - 1, -1, -1):
            row = df.iloc[data_row_index]
            pre_script_value = row.get('preScript')
            ori_step_value = row.get('oriStep') if 'oriStep' in df.columns else None

            if self._has_meaningful_cell_value(pre_script_value) or self._has_meaningful_cell_value(ori_step_value):
                continue

            has_other_content = any(
                self._has_meaningful_cell_value(value)
                for column_name, value in row.items()
                if column_name not in {'preScript', 'oriStep'}
            )
            if has_other_content:
                return data_row_index

        raise ValueError('未找到 oriStep 和 preScript 都为空的可写入数据行')

    def _find_latest_blank_prescript_excel_row(self, worksheet, pre_script_column_index: int, ori_step_column_index: Optional[int] = None) -> int:
        for excel_row in range(worksheet.max_row, 1, -1):
            pre_script_value = worksheet.cell(row=excel_row, column=pre_script_column_index).value
            ori_step_value = worksheet.cell(row=excel_row, column=ori_step_column_index).value if ori_step_column_index else None

            if self._has_meaningful_cell_value(pre_script_value) or self._has_meaningful_cell_value(ori_step_value):
                continue

            has_other_content = any(
                self._has_meaningful_cell_value(worksheet.cell(row=excel_row, column=column_index).value)
                for column_index in range(1, worksheet.max_column + 1)
                if column_index != pre_script_column_index and column_index != ori_step_column_index
            )
            if has_other_content:
                return excel_row

        raise ValueError('未找到 oriStep 和 preScript 都为空的可写入数据行')

    def _build_append_metadata(self, case_number: str, prescript: str = '') -> dict[str, str]:
        normalized_case_number = self._normalize_case_number(case_number)

        # 解析 preScript 中所有 Assert 的 x 值之和，生成对应数量的 (1,1)
        total_assert_count = self._parse_total_assert_count(prescript)
        if total_assert_count > 0:
            check_point = ','.join(['(1,1)'] * total_assert_count)
        else:
            check_point = '(1,1)'

        metadata: dict[str, str] = {
            'runOption': 'Y',
            'original': 'Y',
            'checkPic': self._build_unique_check_pic_name(normalized_case_number),
            'checkPoint': check_point,
        }

        if normalized_case_number:
            metadata.update({
                'testID': normalized_case_number,
                'category': normalized_case_number,
                'testItem': normalized_case_number,
            })

        return metadata

    def _write_append_metadata_with_pandas(self, df: pd.DataFrame, data_row_index: int, case_number: str, prescript: str = '') -> dict[str, str]:
        metadata = self._build_append_metadata(case_number, prescript)
        if not metadata:
            return {}

        for column_name, value in metadata.items():
            if column_name not in df.columns:
                df[column_name] = None
            else:
                df[column_name] = df[column_name].astype(object)
            df.loc[data_row_index, column_name] = value

        return metadata

    def _write_append_metadata_to_sheet(self, worksheet, header_map: Dict[str, int], excel_row: int, case_number: str, prescript: str = '') -> dict[str, str]:
        metadata = self._build_append_metadata(case_number, prescript)
        if not metadata:
            return {}

        for column_name, value in metadata.items():
            column_index = self._ensure_column(worksheet, header_map, column_name)
            self._write_sheet_cell(worksheet, excel_row, column_index, value)

        return metadata

    def _write_cell_with_pandas(self, file_path: Path, file_name: str, column_name: str, row_index: int, value: str) -> Dict[str, Any]:
        df = pd.read_excel(file_path)
        col = column_name
        if col not in df.columns:
            df[col] = None

        ri = int(row_index)
        if ri < 0:
            ri = 0
        if ri >= len(df):
            df = df.reindex(range(ri + 1))

        df.loc[ri, col] = value
        df.to_excel(file_path, index=False)

        return {
            "status": "ok",
            "file": file_name,
            "row_index": ri,
            "column_name": col,
            "preserved_format": False,
        }

    def _update_case_fields_with_pandas(
        self,
        file_path: Path,
        file_name: str,
        excel_row: int,
        title: str,
        ori_step: str,
        pre_script: str,
        verify_image: str,
        step: Optional[str] = None,
        test_result: str = '',
    ) -> Dict[str, Any]:
        df = pd.read_excel(file_path)

        data_row_index = int(excel_row) - 2
        if data_row_index < 0 or data_row_index >= len(df):
            raise ValueError(f"Excel行号超出范围: {excel_row}")

        title_column = 'testID' if 'testID' in df.columns else 'title'
        verify_image_column = 'checkPic' if 'checkPic' in df.columns else 'verify_image'
        has_split_step_columns = 'oriStep' in df.columns or 'preScript' in df.columns

        if title_column not in df.columns:
            df[title_column] = None
        if verify_image_column not in df.columns:
            df[verify_image_column] = None

        df.loc[data_row_index, title_column] = title
        df.loc[data_row_index, verify_image_column] = verify_image

        columns = {
            "title": title_column,
            "verify_image": verify_image_column,
        }

        if test_result is not None and test_result != '':
            test_result_column = 'result' if 'result' in df.columns else 'testResult'
            if test_result_column not in df.columns:
                df[test_result_column] = None
            df.loc[data_row_index, test_result_column] = test_result
            columns["test_result"] = test_result_column

        if has_split_step_columns:
            if 'oriStep' not in df.columns:
                df['oriStep'] = None
            if 'preScript' not in df.columns:
                df['preScript'] = None

            df.loc[data_row_index, 'oriStep'] = ori_step
            df.loc[data_row_index, 'preScript'] = pre_script
            columns.update({
                "ori_step": 'oriStep',
                "pre_script": 'preScript',
            })
        else:
            step_column = 'step' if 'step' in df.columns else 'operation' if 'operation' in df.columns else 'step'
            if step_column not in df.columns:
                df[step_column] = None

            fallback_step = ori_step or step or ''
            df.loc[data_row_index, step_column] = fallback_step
            columns["step"] = step_column

        df.to_excel(file_path, index=False)

        return {
            "status": "ok",
            "file": file_name,
            "excel_row": int(excel_row),
            "data_row_index": data_row_index,
            "columns": columns,
            "preserved_format": False,
        }

    def write_cell(self, file_name: str, column_name: str, row_index: int, value: str) -> Dict[str, Any]:
        """写入单元格"""
        file_path = resolve_excel_file(file_name)
        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_name}")

        if not self._supports_cell_level_write(file_path):
            return self._write_cell_with_pandas(file_path, file_name, column_name, row_index, value)

        last_error = None
        # 优先尝试 openpyxl 写入（保留格式），失败则回退到 pandas
        try:
            workbook, worksheet = self._load_workbook_sheet(file_path)
            header_map = self._get_header_map(worksheet)
            column_index = self._ensure_column(worksheet, header_map, column_name)

            ri = int(row_index)
            if ri < 0:
                ri = 0

            excel_row_index = ri + 2
            self._write_sheet_cell(worksheet, excel_row_index, column_index, value)
            workbook.save(file_path)
            workbook.close()

            return {
                "status": "ok",
                "file": file_name,
                "row_index": ri,
                "column_name": column_name,
                "preserved_format": True,
            }
        except Exception as e:
            last_error = e

        # openpyxl 失败时回退到 pandas
        try:
            print(f"[write_cell] openpyxl 失败({last_error})，回退到 pandas 写入")
            return self._write_cell_with_pandas(file_path, file_name, column_name, row_index, value)
        except Exception as e:
            raise Exception(f"写入失败: openpyxl={last_error}, pandas={e}")

    def append_sequence_to_latest_prescript(
        self,
        file_name: str,
        sequence: str,
        case_number: Optional[str] = None,
        assert_format: Optional[str] = None,
        check_pic: Optional[str] = None,
        check_point: Optional[str] = None,
    ) -> Dict[str, Any]:
        """将序列写入 preScript 列。

        新规则：如果 ``case_number`` 对应的行已经存在 preScript，则在该行末尾追加
        ``,Assert/1/1,<新序列>``（或用户自定义的 assert 格式），并基于 Excel 中已有
        checkPic 生成下一个 -N 版本，而不是再开新行。
        这样同一用例的多次写入会形成连续的步骤+断言序列。
        如果 assert_format 为空字符串，则不插入 Assert。
        如果提供了 check_pic 和 check_point，则直接使用它们，而不是自己生成。
        """
        file_path = resolve_excel_file(file_name)
        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_name}")

        # 直接使用传入的 sequence 作为 preScript（前端已经拼接好了完整的 preScript）
        normalized_sequence = self._compress_adjacent_command_sequence(sequence)
        normalized_case_number = self._normalize_case_number(case_number)
        if not normalized_sequence:
            raise ValueError('待写入的序列不能为空')

        try:
            if not self._supports_cell_level_write(file_path):
                df = pd.read_excel(file_path)

                # 优先尝试在已存在的同 case 行追加 Assert + 新序列
                existing_row_index = self._find_existing_row_for_case_with_pandas(df, normalized_case_number)
                if existing_row_index is not None:
                    if 'preScript' not in df.columns:
                        df['preScript'] = None
                    df['preScript'] = df['preScript'].astype(object)

                    # 直接使用传入的 sequence 作为 preScript（前端已经拼接好了）
                    df.loc[existing_row_index, 'preScript'] = normalized_sequence

                    # 使用传入的 check_pic 和 check_point，或者自己生成
                    if check_pic and check_point:
                        check_pic_column = 'checkPic' if 'checkPic' in df.columns else 'verify_image'
                        if check_pic_column not in df.columns:
                            check_pic_column = 'checkPic'
                            df[check_pic_column] = None
                        df[check_pic_column] = df[check_pic_column].astype(object)
                        df.loc[existing_row_index, check_pic_column] = check_pic

                        check_point_column = 'checkPoint'
                        if check_point_column not in df.columns:
                            df[check_point_column] = None
                        df[check_point_column] = df[check_point_column].astype(object)
                        df.loc[existing_row_index, check_point_column] = check_point

                        # 设置 runOption 和 original
                        for col_name in ('runOption', 'original'):
                            if col_name not in df.columns:
                                df[col_name] = None
                            df[col_name] = df[col_name].astype(object)
                        df.loc[existing_row_index, 'runOption'] = 'Y'
                        df.loc[existing_row_index, 'original'] = 'Y'

                        written_columns = ["preScript", check_pic_column, check_point_column]
                    else:
                        # 自己生成
                        resolved_assert_format = (assert_format or '').strip()
                        total_assert_count = self._parse_total_assert_count(normalized_sequence)
                        if total_assert_count > 0:
                            check_point = ','.join(['(1,1)'] * total_assert_count)
                        else:
                            check_point = '(1,1)'
                        check_point_column = 'checkPoint'
                        if check_point_column not in df.columns:
                            df[check_point_column] = None
                        df[check_point_column] = df[check_point_column].astype(object)
                        df.loc[existing_row_index, check_point_column] = check_point

                        assert_count = self._parse_assert_count(resolved_assert_format)
                        check_pic_column = 'checkPic' if 'checkPic' in df.columns else 'verify_image'
                        if check_pic_column not in df.columns:
                            check_pic_column = 'checkPic'
                            df[check_pic_column] = None
                        df[check_pic_column] = df[check_pic_column].astype(object)
                        old_check_pic = df.loc[existing_row_index, check_pic_column] if check_pic_column in df.columns else ''
                        appended_check_pic = self._build_check_pic_for_append(old_check_pic, normalized_case_number, assert_count)
                        check_pic = appended_check_pic
                        df.loc[existing_row_index, check_pic_column] = appended_check_pic

                        written_columns = ["preScript", check_pic_column]

                    df.to_excel(file_path, index=False)

                    return {
                        "status": "ok",
                        "file": file_name,
                        "excel_row": existing_row_index + 2,
                        "data_row_index": existing_row_index,
                        "column_name": "preScript",
                        "case_number": normalized_case_number,
                        "check_pic": check_pic or '',
                        "check_point": check_point or '(1,1)',
                        "written_columns": written_columns,
                        "case_number_columns": [],
                        "appended_new_row": False,
                        "appended_to_existing_case": True,
                        "preserved_format": False,
                    }

                appended_new_row = False
                try:
                    data_row_index = self._find_latest_blank_prescript_data_row_with_pandas(df)
                except ValueError:
                    data_row_index = len(df)
                    df = df.reindex(range(data_row_index + 1))
                    if 'preScript' not in df.columns:
                        df['preScript'] = None
                    appended_new_row = True

                df['preScript'] = df['preScript'].astype(object)
                # 直接使用传入的 sequence 作为 preScript
                df.loc[data_row_index, 'preScript'] = normalized_sequence

                # 使用传入的 check_pic 和 check_point，或者自己生成
                if check_pic and check_point:
                    # 直接使用传入的值
                    for col_name in ('checkPic', 'checkPoint', 'runOption', 'original'):
                        if col_name not in df.columns:
                            df[col_name] = None
                        df[col_name] = df[col_name].astype(object)
                    df.loc[data_row_index, 'checkPic'] = check_pic
                    df.loc[data_row_index, 'checkPoint'] = check_point
                    df.loc[data_row_index, 'runOption'] = 'Y'
                    df.loc[data_row_index, 'original'] = 'Y'
                    if normalized_case_number:
                        for col_name in ('testID', 'category', 'testItem'):
                            if col_name not in df.columns:
                                df[col_name] = None
                            df[col_name] = df[col_name].astype(object)
                            df.loc[data_row_index, col_name] = normalized_case_number
                    case_number_columns = ['testID', 'category', 'testItem'] if normalized_case_number else []
                    written_columns = ['preScript', 'checkPic', 'checkPoint']
                else:
                    # 自己生成
                    append_metadata = self._write_append_metadata_with_pandas(df, data_row_index, normalized_case_number, normalized_sequence)
                    case_number_columns = [column_name for column_name in ('testID', 'category', 'testItem') if column_name in append_metadata]
                    check_pic = append_metadata.get('checkPic', '')
                    check_point = append_metadata.get('checkPoint', '(1,1)')
                    written_columns = list(append_metadata.keys())

                df.to_excel(file_path, index=False)

                return {
                    "status": "ok",
                    "file": file_name,
                    "excel_row": data_row_index + 2,
                    "data_row_index": data_row_index,
                    "column_name": "preScript",
                    "case_number": normalized_case_number,
                    "check_pic": check_pic or '',
                    "check_point": check_point or '(1,1)',
                    "written_columns": written_columns,
                    "case_number_columns": case_number_columns,
                    "appended_new_row": appended_new_row,
                    "appended_to_existing_case": False,
                    "preserved_format": False,
                }

            workbook, worksheet = self._load_workbook_sheet(file_path)
            try:
                header_map = self._get_header_map(worksheet)
                pre_script_column_index = self._ensure_column(worksheet, header_map, 'preScript')
                ori_step_column_index = header_map.get('oriStep')

                # 优先尝试在已存在的同 case 行追加 Assert + 新序列
                existing_excel_row = self._find_existing_row_for_case_in_sheet(
                    worksheet, header_map, normalized_case_number,
                )
                if existing_excel_row is not None:
                    # 直接使用传入的 sequence 作为 preScript（前端已经拼接好了）
                    self._write_sheet_cell(worksheet, existing_excel_row, pre_script_column_index, normalized_sequence)

                    # 使用传入的 check_pic 和 check_point，或者自己生成
                    if check_pic and check_point:
                        check_pic_column = 'checkPic' if 'checkPic' in header_map else 'verify_image'
                        check_pic_column_index = self._ensure_column(worksheet, header_map, check_pic_column)
                        self._write_sheet_cell(worksheet, existing_excel_row, check_pic_column_index, check_pic)

                        check_point_column_index = self._ensure_column(worksheet, header_map, 'checkPoint')
                        self._write_sheet_cell(worksheet, existing_excel_row, check_point_column_index, check_point)

                        # 设置 runOption 和 original
                        run_option_column_index = self._ensure_column(worksheet, header_map, 'runOption')
                        self._write_sheet_cell(worksheet, existing_excel_row, run_option_column_index, 'Y')
                        original_column_index = self._ensure_column(worksheet, header_map, 'original')
                        self._write_sheet_cell(worksheet, existing_excel_row, original_column_index, 'Y')

                        written_columns = ["preScript", check_pic_column, "checkPoint"]
                    else:
                        # 自己生成
                        resolved_assert_format = (assert_format or '').strip()
                        total_assert_count = self._parse_total_assert_count(normalized_sequence)
                        if total_assert_count > 0:
                            check_point = ','.join(['(1,1)'] * total_assert_count)
                        else:
                            check_point = '(1,1)'
                        check_point_column_index = self._ensure_column(worksheet, header_map, 'checkPoint')
                        self._write_sheet_cell(worksheet, existing_excel_row, check_point_column_index, check_point)

                        assert_count = self._parse_assert_count(resolved_assert_format)
                        check_pic_column = 'checkPic' if 'checkPic' in header_map else 'verify_image'
                        check_pic_column_index = self._ensure_column(worksheet, header_map, check_pic_column)
                        old_check_pic = worksheet.cell(row=existing_excel_row, column=check_pic_column_index).value
                        appended_check_pic = self._build_check_pic_for_append(old_check_pic, normalized_case_number, assert_count)
                        check_pic = appended_check_pic
                        self._write_sheet_cell(worksheet, existing_excel_row, check_pic_column_index, appended_check_pic)

                        written_columns = ["preScript", check_pic_column]

                    workbook.save(file_path)

                    return {
                        "status": "ok",
                        "file": file_name,
                        "excel_row": existing_excel_row,
                        "data_row_index": existing_excel_row - 2,
                        "column_name": "preScript",
                        "case_number": normalized_case_number,
                        "check_pic": check_pic or '',
                        "check_point": check_point or '(1,1)',
                        "written_columns": written_columns,
                        "case_number_columns": [],
                        "appended_new_row": False,
                        "appended_to_existing_case": True,
                        "preserved_format": True,
                    }

                appended_new_row = False
                try:
                    excel_row = self._find_latest_blank_prescript_excel_row(worksheet, pre_script_column_index, ori_step_column_index)
                except ValueError:
                    excel_row = worksheet.max_row + 1
                    appended_new_row = True

                # 直接使用传入的 sequence 作为 preScript
                self._write_sheet_cell(worksheet, excel_row, pre_script_column_index, normalized_sequence)

                # 使用传入的 check_pic 和 check_point，或者自己生成
                if check_pic and check_point:
                    check_pic_column = 'checkPic' if 'checkPic' in header_map else 'verify_image'
                    check_pic_column_index = self._ensure_column(worksheet, header_map, check_pic_column)
                    self._write_sheet_cell(worksheet, excel_row, check_pic_column_index, check_pic)

                    check_point_column_index = self._ensure_column(worksheet, header_map, 'checkPoint')
                    self._write_sheet_cell(worksheet, excel_row, check_point_column_index, check_point)

                    # 设置 runOption 和 original
                    run_option_column_index = self._ensure_column(worksheet, header_map, 'runOption')
                    self._write_sheet_cell(worksheet, excel_row, run_option_column_index, 'Y')
                    original_column_index = self._ensure_column(worksheet, header_map, 'original')
                    self._write_sheet_cell(worksheet, excel_row, original_column_index, 'Y')

                    if normalized_case_number:
                        for col_name in ('testID', 'category', 'testItem'):
                            col_index = self._ensure_column(worksheet, header_map, col_name)
                            self._write_sheet_cell(worksheet, excel_row, col_index, normalized_case_number)

                    case_number_columns = ['testID', 'category', 'testItem'] if normalized_case_number else []
                    written_columns = ['preScript', check_pic_column, 'checkPoint']
                else:
                    # 自己生成
                    append_metadata = self._write_append_metadata_to_sheet(worksheet, header_map, excel_row, normalized_case_number, normalized_sequence)
                    case_number_columns = [column_name for column_name in ('testID', 'category', 'testItem') if column_name in append_metadata]
                    check_pic = append_metadata.get('checkPic', '')
                    check_point = append_metadata.get('checkPoint', '(1,1)')
                    written_columns = list(append_metadata.keys())

                workbook.save(file_path)

                return {
                    "status": "ok",
                    "file": file_name,
                    "excel_row": excel_row,
                    "data_row_index": excel_row - 2,
                    "column_name": "preScript",
                    "case_number": normalized_case_number,
                    "check_pic": check_pic or '',
                    "check_point": check_point or '(1,1)',
                    "written_columns": written_columns,
                    "case_number_columns": case_number_columns,
                    "appended_new_row": appended_new_row,
                    "appended_to_existing_case": False,
                    "preserved_format": True,
                }
            finally:
                workbook.close()
        except Exception as e:
            raise Exception(f"写入 preScript 失败: {e}")

    def update_case_fields(
        self,
        file_name: str,
        excel_row: int,
        title: str,
        ori_step: str,
        pre_script: str,
        verify_image: str,
        step: Optional[str] = None,
        test_result: str = '',
    ) -> Dict[str, Any]:
        """按 Excel 行号更新图片校验用例的展示字段。"""
        file_path = resolve_excel_file(file_name)
        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_name}")

        try:
            if not self._supports_cell_level_write(file_path):
                return self._update_case_fields_with_pandas(
                    file_path,
                    file_name,
                    excel_row,
                    title,
                    ori_step,
                    pre_script,
                    verify_image,
                    step,
                    test_result,
                )

            workbook, worksheet = self._load_workbook_sheet(file_path)
            header_map = self._get_header_map(worksheet)

            data_row_index = int(excel_row) - 2
            actual_excel_row = int(excel_row)
            if data_row_index < 0 or actual_excel_row > worksheet.max_row:
                raise ValueError(f"Excel行号超出范围: {excel_row}")

            title_column = 'testID' if 'testID' in header_map else 'title'
            verify_image_column = 'checkPic' if 'checkPic' in header_map else 'verify_image'
            has_split_step_columns = 'oriStep' in header_map or 'preScript' in header_map

            title_column_index = self._ensure_column(worksheet, header_map, title_column)
            verify_image_column_index = self._ensure_column(worksheet, header_map, verify_image_column)

            self._write_sheet_cell(worksheet, actual_excel_row, title_column_index, title)
            self._write_sheet_cell(worksheet, actual_excel_row, verify_image_column_index, verify_image)

            columns = {
                "title": title_column,
                "verify_image": verify_image_column,
            }

            if test_result is not None and test_result != '':
                test_result_column = 'result' if 'result' in header_map else 'testResult'
                test_result_column_index = self._ensure_column(worksheet, header_map, test_result_column)
                self._write_sheet_cell(worksheet, actual_excel_row, test_result_column_index, test_result)
                columns["test_result"] = test_result_column

            if has_split_step_columns:
                ori_step_column_index = self._ensure_column(worksheet, header_map, 'oriStep')
                pre_script_column_index = self._ensure_column(worksheet, header_map, 'preScript')

                self._write_sheet_cell(worksheet, actual_excel_row, ori_step_column_index, ori_step)
                self._write_sheet_cell(worksheet, actual_excel_row, pre_script_column_index, pre_script)
                columns.update({
                    "ori_step": 'oriStep',
                    "pre_script": 'preScript',
                })
            else:
                step_column = 'step' if 'step' in header_map else 'operation' if 'operation' in header_map else 'step'
                step_column_index = self._ensure_column(worksheet, header_map, step_column)

                fallback_step = ori_step or step or ''
                self._write_sheet_cell(worksheet, actual_excel_row, step_column_index, fallback_step)
                columns["step"] = step_column

            workbook.save(file_path)
            workbook.close()

            return {
                "status": "ok",
                "file": file_name,
                "excel_row": int(excel_row),
                "data_row_index": data_row_index,
                "columns": columns,
                "preserved_format": True,
            }
        except Exception as e:
            raise Exception(f"更新用例字段失败: {e}")

excel_service = ExcelService()
