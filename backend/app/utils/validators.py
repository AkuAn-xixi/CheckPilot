"""Excel验证器模块"""
import openpyxl
from collections import Counter
from typing import Dict, List, Any
from ...FieldValidation import get_valid_keys as get_runtime_valid_keys
from .adb_controller import get_keycode_map, get_custom_commands, NON_EXECUTABLE_KEYS, _parse_key_and_hold, is_valid_repeat_spec
from .path_resolver import resolve_image_file

VALID_KEYS = get_runtime_valid_keys()
ASR_META_COMMANDS = {"TTS"}

class ExcelValidator:
    """Excel文件验证器"""

    @staticmethod
    def validate(file_path: str) -> Dict[str, Any]:
        """验证 Excel 文件的格式和内容"""
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
        except Exception as e:
            return {"success": False, "errors": [f"无法打开Excel文件: {e}"], "warnings": [], "total_rows": 0}

        errors = []
        warnings = []
        valid_keys = {str(key).strip().upper() for key in get_runtime_valid_keys() if str(key).strip()}
        keycode_keys = {
            str(key).strip().upper()
            for key in set(get_keycode_map().keys()) | set(get_custom_commands().keys())
            if str(key).strip()
        }

        c_values = []
        for i in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=i, column=3).value
            if cell_value is not None:
                c_values.append(str(cell_value))

        if len(c_values) != len(set(c_values)):
            duplicates = [item for item, count in Counter(c_values).items() if count > 1]
            errors.append(f"C列值不唯一，重复的值有: {', '.join(duplicates)}")

        for col in ["F", "G"]:
            col_idx = openpyxl.utils.column_index_from_string(col)
            for row in range(2, sheet.max_row + 1):
                cell_value = str(sheet.cell(row=row, column=col_idx).value)
                if cell_value == "None" or not cell_value.strip():
                    continue

                if cell_value.strip().endswith(","):
                    errors.append(f"{col}{row} 结尾有多余的逗号: '{cell_value.strip()}'")

                commands = [cmd.strip() for cmd in cell_value.split(",") if cmd.strip()]
                for cmd in commands:
                    if cmd.upper() in ASR_META_COMMANDS:
                        continue

                    parts = cmd.split("/")
                    if len(parts) != 3:
                        errors.append(f"{col}{row} 命令 '{cmd}' 格式应为 KEY/COUNT/TIME")
                        continue

                    key, _hold_us = _parse_key_and_hold(parts[0])
                    count = parts[1]
                    time_val = parts[2]
                    # NON_EXECUTABLE_KEYS（如 ASSERT）作为系统级占位按键，
                    # 总是视作合法且不要求 keycode 映射
                    if key in NON_EXECUTABLE_KEYS:
                        pass
                    elif key not in valid_keys and key not in ASR_META_COMMANDS:
                        errors.append(f"{col}{row} 按键名称 '{key}' 无效")
                    elif key not in keycode_keys and key not in ASR_META_COMMANDS:
                        errors.append(f"{col}{row} 按键名称 '{key}' 缺少键值映射")

                    if not is_valid_repeat_spec(count):
                        errors.append(f"{col}{row} 按键次数 '{count}' 必须为正整数或 X / X:N / X:(A:B)")

                    try:
                        float(time_val)
                    except ValueError:
                        errors.append(f"{col}{row} 时间参数 '{time_val}' 必须为数字")

        for row in range(2, sheet.max_row + 1):
            cell_value = str(sheet.cell(row=row, column=9).value)
            if cell_value == "None" or not cell_value.strip():
                continue

            image_path_value = cell_value.strip()
            lower_path = image_path_value.lower()
            if not lower_path.endswith((".png", ".jpg", ".jpeg", ".bmp", ".webp")):
                errors.append(f"I{row} 必须为图片路径，支持 png/jpg/jpeg/bmp/webp，当前值: {cell_value}")
                continue

            # resolved_image_path = resolve_image_file(image_path_value, excel_file_path=file_path)
            # if not resolved_image_path.exists():
            #     warnings.append(f"I{row} 图片路径不存在: {image_path_value}")

        import re
        for row in range(2, sheet.max_row + 1):
            cell_value = str(sheet.cell(row=row, column=10).value)
            if cell_value == "None" or not cell_value.strip():
                continue

            pairs = re.findall(r'\([^()]+\)', cell_value.strip())
            if not pairs:
                errors.append(f"J{row} 必须使用英文括号，格式应为(数字,数字)或多个(数字,数字)，当前值: {cell_value}")
                continue

            for pair in pairs:
                inner = pair[1:-1]
                parts = [p.strip() for p in inner.split(",", 1)]
                if len(parts) != 2:
                    errors.append(f"J{row} 格式错误，应为(数字,数字)，当前值: {pair}")
                    break
                try:
                    float(parts[0])
                    float(parts[1])
                except ValueError:
                    errors.append(f"J{row} 包含非数字内容，当前值: {pair}")
                    break

        return {
            "success": len(errors) == 0,
            "errors": errors,
            "warnings": warnings,
            "total_rows": sheet.max_row - 1 if sheet.max_row > 1 else 0
        }
