import tempfile
import unittest
import json
from pathlib import Path
from unittest import mock

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from backend.app.api.customization import KeyCodesUpdateRequest, update_key_codes
from backend.app.config import settings
from backend.app.services.excel_service import ExcelService
from backend.app.utils.adb_controller import ADBController
from backend.app.utils.validators import ExcelValidator


class CompressAdjacentSequenceTests(unittest.TestCase):
    def test_merges_adjacent_numeric_repeats(self):
        self.assertEqual(
            ExcelService._compress_adjacent_command_sequence("HOME/1/1,HOME/1/1"),
            "HOME/2/1",
        )

    def test_preserves_random_repeat_as_is(self):
        self.assertEqual(
            ExcelService._compress_adjacent_command_sequence("OK/X:5/1,DOWN/X/3"),
            "OK/X:5/1,DOWN/X/3",
        )

    def test_numeric_part_does_not_merge_into_random_part(self):
        self.assertEqual(
            ExcelService._compress_adjacent_command_sequence("OK/X:5/1,OK/1/1,OK/2/1"),
            "OK/X:5/1,OK/3/1",
        )

    def test_random_part_does_not_merge_into_previous_numeric_part(self):
        self.assertEqual(
            ExcelService._compress_adjacent_command_sequence("OK/1/1,OK/X:3/1,OK/1/1"),
            "OK/1/1,OK/X:3/1,OK/1/1",
        )

    def test_lowercase_x_also_preserved(self):
        self.assertEqual(
            ExcelService._compress_adjacent_command_sequence("OK/x/1,OK/2/1"),
            "OK/x/1,OK/2/1",
        )


class UpdateCaseFieldsTests(unittest.TestCase):
    def setUp(self):
        self.service = ExcelService()

    def test_update_case_fields_updates_split_step_columns(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            excel_path = Path(tmp_dir) / 'cases.xlsx'
            pd.DataFrame([
                {
                    'testID': 'Old Title',
                    'oriStep': 'HOME/1/1',
                    'preScript': 'LEFT/1/1',
                    'checkPic': 'old.png',
                }
            ]).to_excel(excel_path, index=False)

            with mock.patch('backend.app.services.excel_service.resolve_excel_file', return_value=excel_path):
                result = self.service.update_case_fields(
                    'cases.xlsx',
                    2,
                    'New Title',
                    'OK/1/1',
                    'DOWN/1/1',
                    'new.png',
                )

            updated_df = pd.read_excel(excel_path)
            self.assertEqual(updated_df.loc[0, 'testID'], 'New Title')
            self.assertEqual(updated_df.loc[0, 'oriStep'], 'OK/1/1')
            self.assertEqual(updated_df.loc[0, 'preScript'], 'DOWN/1/1')
            self.assertEqual(updated_df.loc[0, 'checkPic'], 'new.png')
            self.assertEqual(result['columns']['ori_step'], 'oriStep')
            self.assertEqual(result['columns']['pre_script'], 'preScript')

    def test_update_case_fields_falls_back_to_single_step_column(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            excel_path = Path(tmp_dir) / 'legacy.xlsx'
            pd.DataFrame([
                {
                    'title': 'Legacy Title',
                    'step': 'HOME/1/1',
                    'verify_image': 'old.png',
                }
            ]).to_excel(excel_path, index=False)

            with mock.patch('backend.app.services.excel_service.resolve_excel_file', return_value=excel_path):
                result = self.service.update_case_fields(
                    'legacy.xlsx',
                    2,
                    'Legacy Updated',
                    'BACK/1/1',
                    'IGNORED/1/1',
                    'new.png',
                )

            updated_df = pd.read_excel(excel_path)
            self.assertEqual(updated_df.loc[0, 'title'], 'Legacy Updated')
            self.assertEqual(updated_df.loc[0, 'step'], 'BACK/1/1')
            self.assertEqual(updated_df.loc[0, 'verify_image'], 'new.png')
            self.assertEqual(result['columns']['step'], 'step')

    def test_update_case_fields_preserves_existing_cell_style(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            excel_path = Path(tmp_dir) / 'styled.xlsx'
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(['testID', 'oriStep', 'preScript', 'checkPic'])
            worksheet.append(['Old Title', 'HOME/1/1', 'LEFT/1/1', 'old.png'])
            worksheet['A2'].fill = PatternFill(fill_type='solid', fgColor='FFFF00')
            worksheet['A2'].font = Font(bold=True, color='FF0000')
            worksheet['D2'].fill = PatternFill(fill_type='solid', fgColor='00FF00')
            workbook.save(excel_path)
            workbook.close()

            with mock.patch('backend.app.services.excel_service.resolve_excel_file', return_value=excel_path):
                self.service.update_case_fields(
                    'styled.xlsx',
                    2,
                    'Styled Title',
                    'OK/1/1',
                    'DOWN/1/1',
                    'styled.png',
                )

            updated_workbook = load_workbook(excel_path)
            updated_worksheet = updated_workbook.active
            self.assertEqual(updated_worksheet['A2'].value, 'Styled Title')
            self.assertEqual(updated_worksheet['D2'].value, 'styled.png')
            self.assertEqual(updated_worksheet['A2'].fill.fill_type, 'solid')
            self.assertTrue((updated_worksheet['A2'].fill.fgColor.rgb or '').endswith('FFFF00'))
            self.assertTrue(updated_worksheet['A2'].font.bold)
            self.assertTrue((updated_worksheet['D2'].fill.fgColor.rgb or '').endswith('00FF00'))
            updated_workbook.close()


class WriteCellTests(unittest.TestCase):
    def setUp(self):
        self.service = ExcelService()

    def test_write_cell_preserves_existing_cell_style(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            excel_path = Path(tmp_dir) / 'sequence.xlsx'
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(['sequence'])
            worksheet.append(['old'])
            worksheet['A2'].fill = PatternFill(fill_type='solid', fgColor='FFCC00')
            worksheet['A2'].font = Font(italic=True)
            workbook.save(excel_path)
            workbook.close()

            with mock.patch('backend.app.services.excel_service.resolve_excel_file', return_value=excel_path):
                result = self.service.write_cell('sequence.xlsx', 'sequence', 0, 'new')

            updated_workbook = load_workbook(excel_path)
            updated_worksheet = updated_workbook.active
            self.assertEqual(updated_worksheet['A2'].value, 'new')
            self.assertEqual(updated_worksheet['A2'].fill.fill_type, 'solid')
            self.assertTrue((updated_worksheet['A2'].fill.fgColor.rgb or '').endswith('FFCC00'))
            self.assertTrue(updated_worksheet['A2'].font.italic)
            self.assertEqual(result['column_name'], 'sequence')
            updated_workbook.close()


class AppendSequenceTests(unittest.TestCase):
    def setUp(self):
        self.service = ExcelService()

    def test_append_sequence_writes_case_number_to_case_columns_with_pandas(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            excel_path = Path(tmp_dir) / 'cases.xlsx'
            test_cases_dir = Path(tmp_dir) / 'test_cases'
            pd.DataFrame([
                {
                    'title': 'Placeholder Row',
                    'oriStep': None,
                    'preScript': None,
                }
            ]).to_excel(excel_path, index=False)

            with mock.patch('backend.app.services.excel_service.resolve_excel_file', return_value=excel_path), mock.patch.object(
                self.service,
                '_supports_cell_level_write',
                return_value=False,
            ), mock.patch.object(settings, 'TEST_CASES_DIR', test_cases_dir):
                result = self.service.append_sequence_to_latest_prescript('cases.xlsx', 'HOME/1/1', 'CASE-9001')

            updated_df = pd.read_excel(excel_path)
            self.assertEqual(updated_df.loc[0, 'preScript'], 'HOME/1/1')
            self.assertEqual(updated_df.loc[0, 'testID'], 'CASE-9001')
            self.assertEqual(updated_df.loc[0, 'category'], 'CASE-9001')
            self.assertEqual(updated_df.loc[0, 'testItem'], 'CASE-9001')
            self.assertEqual(updated_df.loc[0, 'runOption'], 'Y')
            self.assertEqual(updated_df.loc[0, 'original'], 'Y')
            self.assertEqual(updated_df.loc[0, 'checkPic'], 'CASE-9001.png')
            self.assertEqual(result['check_pic'], 'CASE-9001.png')
            self.assertEqual(
                result['written_columns'],
                ['runOption', 'original', 'checkPic', 'checkPoint', 'testID', 'category', 'testItem'],
            )
            self.assertEqual(result['case_number_columns'], ['testID', 'category', 'testItem'])

    def test_append_sequence_writes_case_number_to_case_columns_with_openpyxl(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            excel_path = Path(tmp_dir) / 'cases.xlsx'
            test_cases_dir = Path(tmp_dir) / 'test_cases'
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(['title', 'oriStep', 'preScript'])
            worksheet.append(['Placeholder Row', None, None])
            workbook.save(excel_path)
            workbook.close()

            with mock.patch('backend.app.services.excel_service.resolve_excel_file', return_value=excel_path), mock.patch.object(settings, 'TEST_CASES_DIR', test_cases_dir):
                result = self.service.append_sequence_to_latest_prescript('cases.xlsx', 'OK/1/1', 'CASE-1001')

            updated_workbook = load_workbook(excel_path)
            updated_worksheet = updated_workbook.active
            self.assertEqual(updated_worksheet['C2'].value, 'OK/1/1')
            self.assertEqual(updated_worksheet['D1'].value, 'runOption')
            self.assertEqual(updated_worksheet['E1'].value, 'original')
            self.assertEqual(updated_worksheet['F1'].value, 'checkPic')
            self.assertEqual(updated_worksheet['G1'].value, 'checkPoint')
            self.assertEqual(updated_worksheet['H1'].value, 'testID')
            self.assertEqual(updated_worksheet['I1'].value, 'category')
            self.assertEqual(updated_worksheet['J1'].value, 'testItem')
            self.assertEqual(updated_worksheet['D2'].value, 'Y')
            self.assertEqual(updated_worksheet['E2'].value, 'Y')
            self.assertEqual(updated_worksheet['F2'].value, 'CASE-1001.png')
            self.assertEqual(updated_worksheet['G2'].value, '(1,1)')
            self.assertEqual(updated_worksheet['H2'].value, 'CASE-1001')
            self.assertEqual(updated_worksheet['I2'].value, 'CASE-1001')
            self.assertEqual(updated_worksheet['J2'].value, 'CASE-1001')
            self.assertEqual(result['check_pic'], 'CASE-1001.png')
            self.assertEqual(
                result['written_columns'],
                ['runOption', 'original', 'checkPic', 'checkPoint', 'testID', 'category', 'testItem'],
            )
            self.assertEqual(result['case_number_columns'], ['testID', 'category', 'testItem'])
            updated_workbook.close()

    def test_build_unique_check_pic_name_returns_base_name_regardless_of_disk(self):
        # 新语义：第一次为某 case 写入时直接给 <case>.png，不再做磁盘冲突避让；
        # 同 case 后续写入由 _build_next_check_pic_name 在 Excel 里递增 -N。
        with tempfile.TemporaryDirectory() as tmp_dir:
            test_cases_dir = Path(tmp_dir) / 'test_cases'
            image_dir = test_cases_dir / 'images'
            image_dir.mkdir(parents=True, exist_ok=True)
            # 故意在磁盘上放同名残留文件，期望仍然返回 base 名
            (image_dir / 'CASE-42.png').write_bytes(b'1')
            (image_dir / 'CASE-42-1.png').write_bytes(b'2')

            with mock.patch.object(settings, 'TEST_CASES_DIR', test_cases_dir):
                file_name = self.service._build_unique_check_pic_name('CASE-42')

            self.assertEqual(file_name, 'CASE-42.png')

    def test_build_next_check_pic_name_increments_numeric_suffix(self):
        # 同 case 的二次写入：从 base.png 进到 base-1.png；再次进到 base-2.png
        self.assertEqual(
            self.service._build_next_check_pic_name('CASE-42.png', 'CASE-42'),
            'CASE-42-1.png',
        )
        self.assertEqual(
            self.service._build_next_check_pic_name('CASE-42-1.png', 'CASE-42'),
            'CASE-42-2.png',
        )
        self.assertEqual(
            self.service._build_next_check_pic_name('CASE-42-7.png', 'CASE-42'),
            'CASE-42-8.png',
        )
        # existing 为空：从 -1 开始
        self.assertEqual(
            self.service._build_next_check_pic_name('', 'CASE-42'),
            'CASE-42-1.png',
        )
        # existing 是逗号分隔的累加值，取最后一张作为参照
        self.assertEqual(
            self.service._build_next_check_pic_name(
                'OSF10_ADV_0002.png,OSF10_ADV_0002-1.png',
                'OSF10_ADV_0002',
            ),
            'OSF10_ADV_0002-2.png',
        )

    def test_append_check_pic_value_concatenates_with_comma(self):
        self.assertEqual(
            self.service._append_check_pic_value('A.png', 'B.png'),
            'A.png,B.png',
        )
        self.assertEqual(
            self.service._append_check_pic_value('A.png,B.png', 'C.png'),
            'A.png,B.png,C.png',
        )
        # 空值：保持另一边
        self.assertEqual(self.service._append_check_pic_value('', 'B.png'), 'B.png')
        self.assertEqual(self.service._append_check_pic_value('A.png', ''), 'A.png')

    def test_append_check_pic_value_strips_legacy_paths(self):
        # 旧版本错误地把绝对路径写进 checkPic 列时，下次拼接前会被剥成 basename
        self.assertEqual(
            self.service._append_check_pic_value(
                r'D:\Checkphoto\OSF10_ADV_0001.png,D:\Checkphoto\OSF10_ADV_0001-1.png',
                'OSF10_ADV_0001-2.png',
            ),
            'OSF10_ADV_0001.png,OSF10_ADV_0001-1.png,OSF10_ADV_0001-2.png',
        )
        # 同时支持 unix 风格路径
        self.assertEqual(
            self.service._append_check_pic_value(
                '/var/images/A.png,/var/images/B.png',
                'C.png',
            ),
            'A.png,B.png,C.png',
        )


class ExcelValidationTests(unittest.TestCase):
    def test_validate_allows_tts_marker_in_command_cells(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            excel_path = Path(tmp_dir) / 'asr.xlsx'
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'])
            worksheet.append([None, None, 'case-1', None, None, 'HOME/1/1,TTS,OK/1/1', None, None, None, None])
            workbook.save(excel_path)
            workbook.close()

            result = ExcelValidator.validate(str(excel_path))

        self.assertTrue(result['success'], result['errors'])

    @mock.patch('backend.app.utils.validators.get_runtime_valid_keys', return_value={'HOME', 'ZEPHYR'})
    @mock.patch('backend.app.utils.validators.get_keycode_map', return_value={'HOME': 3, 'ZEPHYR': 901})
    def test_validate_uses_active_customized_valid_keys(self, _mock_keycodes, _mock_valid_keys):
        with tempfile.TemporaryDirectory() as tmp_dir:
            excel_path = Path(tmp_dir) / 'customized.xlsx'
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'])
            worksheet.append([None, None, 'case-1', None, None, 'ZEPHYR/1/1', None, None, None, None])
            workbook.save(excel_path)
            workbook.close()

            result = ExcelValidator.validate(str(excel_path))

        self.assertTrue(result['success'], result['errors'])

    @mock.patch('backend.app.utils.validators.get_runtime_valid_keys', return_value={'HOME', 'ZEPHYR'})
    @mock.patch('backend.app.utils.validators.get_keycode_map', return_value={'HOME': 3})
    def test_validate_reports_missing_keycode_mapping_for_custom_key(self, _mock_keycodes, _mock_valid_keys):
        with tempfile.TemporaryDirectory() as tmp_dir:
            excel_path = Path(tmp_dir) / 'customized.xlsx'
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'])
            worksheet.append([None, None, 'case-1', None, None, 'ZEPHYR/1/1', None, None, None, None])
            workbook.save(excel_path)
            workbook.close()

            result = ExcelValidator.validate(str(excel_path))

        self.assertFalse(result['success'])
        self.assertIn("按键名称 'ZEPHYR' 缺少键值映射", result['errors'][0])

    @mock.patch('backend.app.utils.validators.get_runtime_valid_keys', return_value={'HOME', 'CLEARNETFLIX'})
    @mock.patch('backend.app.utils.validators.get_custom_commands', return_value={'CLEARNETFLIX': 'adb shell am force-stop com.netflix.ninja'})
    @mock.patch('backend.app.utils.validators.get_keycode_map', return_value={'HOME': 3})
    def test_validate_accepts_custom_command_key(self, _mock_keycodes, _mock_commands, _mock_valid_keys):
        with tempfile.TemporaryDirectory() as tmp_dir:
            excel_path = Path(tmp_dir) / 'custom_commands.xlsx'
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'])
            worksheet.append([None, None, 'case-1', None, None, 'CLEARNETFLIX/1/1', None, None, None, None])
            workbook.save(excel_path)
            workbook.close()

            result = ExcelValidator.validate(str(excel_path))

        self.assertTrue(result['success'], result['errors'])


class CustomizedKeyExecutionTests(unittest.TestCase):
    def test_read_excel_commands_reports_missing_mapping_for_custom_valid_key(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            excel_path = Path(tmp_dir) / 'cases.xlsx'
            pd.DataFrame([
                {'runOption': 'Y', 'oriStep': 'ZEPHYR/1/1', 'preScript': None}
            ]).to_excel(excel_path, index=False)

            with mock.patch('backend.app.utils.adb_controller.get_runtime_valid_keys', return_value={'ZEPHYR'}), mock.patch(
                'backend.app.utils.adb_controller.get_keycode_map',
                return_value={'HOME': 3},
            ):
                result = ADBController().read_excel_commands(str(excel_path))

        self.assertEqual(result['total_rows'], 0)
        self.assertEqual(result['skipped_rows'][0]['reason'], '按键缺少键值映射: ZEPHYR')

    def test_update_key_codes_adds_new_key_to_valid_keys(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            customization_path = Path(tmp_dir) / 'customization.json'
            customization_path.write_text(
                json.dumps({
                    'active_scheme': '默认方案',
                    'schemes': {'默认方案': {}}
                }, ensure_ascii=False, indent=2),
                encoding='utf-8',
            )

            with mock.patch.object(settings, 'CUSTOMIZATION_FILE', customization_path):
                result = update_key_codes('默认方案', KeyCodesUpdateRequest(key_codes={'ZEPHYR': 901}))

            saved = json.loads(customization_path.read_text(encoding='utf-8'))
            scheme = saved['schemes']['默认方案']

        self.assertEqual(result['custom_overrides']['ZEPHYR'], 901)
        self.assertEqual(scheme['key_codes']['ZEPHYR'], 901)
        self.assertIn('ZEPHYR', scheme['valid_keys'])


if __name__ == '__main__':
    unittest.main()